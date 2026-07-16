from __future__ import annotations

import importlib
import json
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from server.services import excel_import_service
from server.services.excel_import_service import (
    IMPORT_COLUMNS,
    build_import_template,
    load_preview,
    preview_import,
)


REQUIRED_COLUMNS = ("供应商内容编号", "活动主题", "平台", "标题", "正文")
NEW_CONTENT_COLUMNS = ("标题", "内容", "类型", "目标平台", "作者", "发布日期")
OPTIMIZED_CONTENT_COLUMNS = ("标题", "内容", "类型", "目标平台", "作者", "发布日期", "优化后版本")


def write_workbook(
    path: Path,
    rows: list[list[object]],
    headers: tuple[str, ...] = IMPORT_COLUMNS,
    *,
    second_sheet_rows: list[list[object]] | None = None,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "内容清单"
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(row)
    if second_sheet_rows is not None:
        second = workbook.create_sheet("ignored")
        second.append(list(headers))
        for row in second_sheet_rows:
            second.append(row)
    workbook.save(path)
    return path


def valid_row(
    external_id: object = "content-1",
    *,
    publish_time: object = None,
    title: object = " 标题 ",
    body: object = " 正文 ",
) -> list[object]:
    return [external_id, " 活动主题 ", " 小红书 ", title, body, publish_time, " 备注 "]


def write_zip(path: Path, entries: list[tuple[object, ...]]) -> Path:
    with ZipFile(path, "w") as archive:
        for entry in entries:
            name, content, *kind = entry
            archive.writestr(name, content, compress_type=kind[0] if kind else ZIP_DEFLATED)
    return path


def assert_has_error(preview, text: str, row_index: int = 0) -> None:
    assert any(text in error for error in preview.rows[row_index].errors), preview.rows[row_index].errors


@pytest.fixture(autouse=True)
def isolate_preview_root_registry(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv(
        "CONTENT_REVIEW_PREVIEW_ROOT_REGISTRY",
        str(tmp_path / "preview-roots.json"),
    )


def test_default_preview_root_registry_path_is_not_cwd_dependent(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.delenv("CONTENT_REVIEW_PREVIEW_ROOT_REGISTRY", raising=False)
    first_cwd = tmp_path / "first"
    second_cwd = tmp_path / "second"
    first_cwd.mkdir()
    second_cwd.mkdir()

    monkeypatch.chdir(first_cwd)
    first_path = excel_import_service._preview_root_registry_path()
    monkeypatch.chdir(second_cwd)

    assert excel_import_service._preview_root_registry_path() == first_path
    assert first_path.is_absolute()


def test_build_import_template_has_exact_chinese_columns() -> None:
    workbook = load_workbook(BytesIO(build_import_template()), read_only=True)

    assert workbook.sheetnames == ["内容清单", "字段说明"]
    assert tuple(cell.value for cell in next(workbook["内容清单"].iter_rows())) == NEW_CONTENT_COLUMNS


def test_preview_accepts_exact_new_headers_maps_fields_and_derives_stable_id(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "new-template.xlsx",
        [[" 新标题 ", " 新内容 ", " 图文 ", " 小红书 ", " 地图作者 ", date(2026, 7, 20)]],
        NEW_CONTENT_COLUMNS,
    )

    first = preview_import(xlsx, None, tmp_path / "first")
    second = preview_import(xlsx, None, tmp_path / "second")

    assert first.valid_count == 1
    assert first.error_count == 0
    assert first.rows[0].normalized == {
        "external_id": second.rows[0].normalized["external_id"],
        "campaign_theme": None,
        "account_name": "地图作者",
        "account_type": "图文",
        "platform": "小红书",
        "title": "新标题",
        "body": "新内容",
        "publish_time": "2026-07-20",
        "note": None,
    }
    assert first.rows[0].normalized["external_id"].startswith("excel:")


def test_new_format_identical_title_and_author_rows_receive_distinct_ids(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "new-duplicate-input.xlsx",
        [
            ["同标题", "内容一", "图文", "微博", "同作者", None],
            ["同标题", "内容二", "图文", "微博", "同作者", None],
        ],
        NEW_CONTENT_COLUMNS,
    )

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.valid_count == 2
    assert len({row.normalized["external_id"] for row in preview.rows}) == 2


def test_preview_accepts_actual_project_headers_with_optimized_body(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "optimized-template.xlsx",
        [["标题", "原始内容", "UGC", "小红书", "作者", "46293", "优化后内容"]],
        OPTIMIZED_CONTENT_COLUMNS,
    )

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.valid_count == 1
    row = preview.rows[0].normalized
    assert row["body"] == "优化后内容"
    assert row["publish_time"] == "2026-09-28"
    assert row["note"] == "导入时使用“优化后版本”作为审核正文"


def test_preview_rejects_duplicate_derived_ids_within_batch(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(excel_import_service, "_derive_external_id", lambda *_: "excel:collision")
    xlsx = write_workbook(
        tmp_path / "derived-collision.xlsx",
        [
            ["标题一", "内容一", "图文", "微博", "作者一", None],
            ["标题二", "内容二", "图文", "微博", "作者二", None],
        ],
        NEW_CONTENT_COLUMNS,
    )

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.valid_count == 0
    assert all(any("内容编号在批次内重复" in error for error in row.errors) for row in preview.rows)


def test_preview_parses_only_first_worksheet_and_trims_required_values(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "input.xlsx",
        [valid_row()],
        second_sheet_rows=[valid_row("ignored")],
    )

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.total_count == 1
    assert preview.valid_count == 1
    assert preview.error_count == 0
    assert preview.rows[0].row_number == 2
    assert preview.rows[0].valid is True
    assert preview.rows[0].normalized == {
        "external_id": "content-1",
        "campaign_theme": "活动主题",
        "platform": "小红书",
        "title": "标题",
        "body": "正文",
        "publish_time": None,
        "note": "备注",
    }


@pytest.mark.parametrize("missing_header", REQUIRED_COLUMNS)
def test_preview_rejects_missing_required_headers(tmp_path: Path, missing_header: str) -> None:
    headers = tuple(header for header in IMPORT_COLUMNS if header != missing_header)
    xlsx = write_workbook(tmp_path / "missing.xlsx", [], headers)

    with pytest.raises(ValueError, match=missing_header):
        preview_import(xlsx, None, tmp_path / "imports")


def test_preview_accepts_workbook_with_only_required_headers(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "required-only.xlsx",
        [valid_row()[: len(REQUIRED_COLUMNS)]],
        REQUIRED_COLUMNS,
    )

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.total_count == 1
    assert preview.valid_count == 1
    assert preview.rows[0].normalized["publish_time"] is None
    assert preview.rows[0].normalized["note"] is None


def test_preview_rejects_blank_and_duplicate_headers(tmp_path: Path) -> None:
    blank = list(IMPORT_COLUMNS)
    blank[-1] = ""
    duplicate = list(IMPORT_COLUMNS)
    duplicate[-1] = "标题"

    with pytest.raises(ValueError, match="表头.*空"):
        preview_import(write_workbook(tmp_path / "blank.xlsx", [], tuple(blank)), None, tmp_path / "a")
    with pytest.raises(ValueError, match="重复.*标题"):
        preview_import(write_workbook(tmp_path / "duplicate.xlsx", [], tuple(duplicate)), None, tmp_path / "b")


def test_preview_marks_duplicate_ids_on_every_affected_row(tmp_path: Path) -> None:
    xlsx = write_workbook(tmp_path / "duplicate-id.xlsx", [valid_row("same"), valid_row(" same ")])

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.error_count == 2
    assert_has_error(preview, "重复")
    assert_has_error(preview, "重复", 1)


def test_preview_ignores_blank_rows_and_rejects_more_than_500_nonblank_rows(tmp_path: Path) -> None:
    rows = [valid_row(f"id-{index}") for index in range(500)]
    rows.insert(20, [None] * len(IMPORT_COLUMNS))
    allowed = preview_import(write_workbook(tmp_path / "500.xlsx", rows), None, tmp_path / "allowed")
    assert allowed.total_count == 500

    rows.append(valid_row("overflow"))
    with pytest.raises(ValueError, match="500"):
        preview_import(write_workbook(tmp_path / "501.xlsx", rows), None, tmp_path / "rejected")


def test_blank_external_id_invalid_row_survives_manifest_round_trip(tmp_path: Path) -> None:
    xlsx = write_workbook(tmp_path / "blank-id.xlsx", [valid_row(" ")])

    preview = preview_import(xlsx, None, tmp_path / "imports")
    loaded = load_preview(preview.token)

    assert loaded == preview
    assert loaded.rows[0].normalized["external_id"] is None
    assert loaded.rows[0].valid is False
    assert_has_error(loaded, "供应商内容编号")


def test_blank_test_owner_does_not_affect_content_precheck(tmp_path: Path) -> None:
    xlsx = write_named_workbook(
        tmp_path / "blank-owner.xlsx",
        [tech_row(" ")],
        [[" ", "case-1", "通过", "指令", "结果", None, None, None, None, None, None, None]],
    )

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.rows[0].tests == []
    assert preview.test_cases == []
    assert preview.errors == []


def test_preview_validates_required_values_and_content_service_lengths(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "invalid-values.xlsx",
        [
            valid_row(" "),
            valid_row("long-title", title="x" * 501),
            valid_row("long-body", body="x" * 100_001),
        ],
    )

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.error_count == 3
    assert_has_error(preview, "供应商内容编号")
    assert_has_error(preview, "标题", 1)
    assert_has_error(preview, "正文", 2)


@pytest.mark.parametrize(
    ("raw_value", "expected"),
    [
        (date(2026, 7, 14), "2026-07-14"),
        (datetime(2026, 7, 15, 13, 30), "2026-07-15"),
        (" 2026-07-16 ", "2026-07-16"),
    ],
)
def test_preview_normalizes_excel_and_text_dates(tmp_path: Path, raw_value: object, expected: str) -> None:
    xlsx = write_workbook(tmp_path / "date.xlsx", [valid_row(publish_time=raw_value)])

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert preview.rows[0].normalized["publish_time"] == expected
    assert preview.rows[0].valid is True


def test_preview_marks_invalid_date_as_row_error(tmp_path: Path) -> None:
    xlsx = write_workbook(tmp_path / "bad-date.xlsx", [valid_row(publish_time="2026/07/14")])

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert_has_error(preview, "计划发布时间")


def test_preview_rejects_zip_in_text_only_mode(tmp_path: Path) -> None:
    xlsx = write_workbook(tmp_path / "text-only.xlsx", [valid_row()])
    archive = write_zip(tmp_path / "media.zip", [("cover.jpg", b"image")])

    with pytest.raises(ValueError, match="仅支持文字"):
        preview_import(xlsx, archive, tmp_path / "imports")


def test_preview_token_is_opaque_and_metadata_is_persisted_without_raw_file_bytes(tmp_path: Path) -> None:
    raw_marker = b"unique raw workbook marker"
    xlsx = write_workbook(tmp_path / "token.xlsx", [valid_row(body=raw_marker.decode())])
    temp_root = tmp_path / "imports"

    preview = preview_import(xlsx, None, temp_root)
    loaded = load_preview(preview.token)

    assert loaded == preview
    assert len(preview.token) >= 32
    assert str(temp_root) not in preview.token
    manifests = list(temp_root.rglob("preview.json"))
    assert len(manifests) == 1
    metadata_bytes = manifests[0].read_bytes()
    json.loads(metadata_bytes)
    assert xlsx.read_bytes() not in metadata_bytes


def _preview_manifest(temp_root: Path, token: str) -> Path:
    return temp_root / token / "preview.json"


def test_preview_survives_module_registry_reset_and_reload(tmp_path: Path, monkeypatch) -> None:
    registry = tmp_path / "preview-roots.json"
    monkeypatch.setenv("CONTENT_REVIEW_PREVIEW_ROOT_REGISTRY", str(registry))
    xlsx = write_workbook(tmp_path / "restart.xlsx", [valid_row()])
    temp_root = tmp_path / "imports"
    preview = preview_import(xlsx, None, temp_root)

    excel_import_service._preview_locations.clear()
    reloaded = importlib.reload(excel_import_service)

    loaded = reloaded.load_preview(preview.token)
    assert reloaded._preview_to_dict(loaded) == reloaded._preview_to_dict(preview)
    registry_payload = json.loads(registry.read_text(encoding="utf-8"))
    assert registry_payload == {"roots": [str(temp_root.resolve())], "version": 1}
    assert str(temp_root) not in preview.token


def test_consume_preview_survives_module_registry_reset_and_reload(tmp_path: Path, monkeypatch) -> None:
    registry = tmp_path / "preview-roots.json"
    monkeypatch.setenv("CONTENT_REVIEW_PREVIEW_ROOT_REGISTRY", str(registry))
    xlsx = write_workbook(tmp_path / "restart-consume.xlsx", [valid_row()])
    temp_root = tmp_path / "imports"
    preview = preview_import(xlsx, None, temp_root)
    preview_dir = temp_root / preview.token

    excel_import_service._preview_locations.clear()
    reloaded = importlib.reload(excel_import_service)
    consumed = reloaded.consume_preview(preview.token)

    assert reloaded._preview_to_dict(consumed) == reloaded._preview_to_dict(preview)
    assert not preview_dir.exists()
    with pytest.raises(ValueError, match="不存在|失效"):
        reloaded.consume_preview(preview.token)


def test_expired_preview_is_rejected_and_cleaned_up(tmp_path: Path) -> None:
    xlsx = write_workbook(tmp_path / "expired.xlsx", [valid_row()])
    temp_root = tmp_path / "imports"
    preview = preview_import(xlsx, None, temp_root)
    preview_dir = temp_root / preview.token
    manifest = _preview_manifest(temp_root, preview.token)
    payload = json.loads(manifest.read_text(encoding="utf-8"))
    payload["expires_at"] = (datetime.now(timezone.utc) - timedelta(seconds=1)).isoformat()
    manifest.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(ValueError, match="过期|失效"):
        load_preview(preview.token)

    assert not preview_dir.exists()


def test_consume_preview_returns_once_and_removes_persisted_preview(tmp_path: Path) -> None:
    xlsx = write_workbook(tmp_path / "consume.xlsx", [valid_row()])
    temp_root = tmp_path / "imports"
    preview = preview_import(xlsx, None, temp_root)
    preview_dir = temp_root / preview.token

    consumed = excel_import_service.consume_preview(preview.token)

    assert consumed == preview
    assert not preview_dir.exists()
    with pytest.raises(ValueError, match="不存在|失效"):
        excel_import_service.consume_preview(preview.token)


def test_formula_cells_are_reported_with_readable_row_and_field_errors(tmp_path: Path) -> None:
    xlsx = write_workbook(tmp_path / "formula.xlsx", [valid_row(title="=1+1")])

    preview = preview_import(xlsx, None, tmp_path / "imports")

    assert_has_error(preview, "第 2 行")
    assert_has_error(preview, "标题")
    assert_has_error(preview, "公式")


def test_formula_only_rows_still_trigger_500_row_limit(tmp_path: Path) -> None:
    formula_rows = [[None] * (len(IMPORT_COLUMNS) - 1) + ["=ROW()"] for _ in range(501)]
    xlsx = write_workbook(tmp_path / "formula-overflow.xlsx", formula_rows)
    temp_root = tmp_path / "imports"

    with pytest.raises(ValueError, match="500"):
        preview_import(xlsx, None, temp_root)

    assert list(temp_root.iterdir()) == []


@pytest.mark.parametrize(
    "mutate",
    [
        lambda payload: payload.update(total_count=99),
        lambda payload: payload["rows"][0].update(valid=False),
        lambda payload: payload["rows"][0]["normalized"].update(unexpected="value"),
        lambda payload: payload["rows"][0]["normalized"].update(external_id=None),
        lambda payload: payload["rows"][0].update(row_number="2"),
        lambda payload: payload.update(unexpected="value"),
        lambda payload: payload.update(expires_at="not-a-timestamp"),
    ],
)
def test_load_preview_rejects_tampered_manifest_schema_and_counts(tmp_path: Path, mutate) -> None:
    xlsx = write_workbook(tmp_path / "tampered.xlsx", [valid_row()])
    temp_root = tmp_path / "imports"
    preview = preview_import(xlsx, None, temp_root)
    manifest = _preview_manifest(temp_root, preview.token)
    payload = json.loads(manifest.read_text(encoding="utf-8"))
    mutate(payload)
    manifest.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(ValueError, match="预览数据无效"):
        load_preview(preview.token)


class _LazyWorksheet:
    def iter_rows(self, *, values_only: bool = False):
        if values_only:
            yield IMPORT_COLUMNS
        else:
            yield tuple(_FormulaAwareCell(value, "s", f"A{index}") for index, value in enumerate(IMPORT_COLUMNS, 1))
        raise BadZipFile("broken worksheet XML")


class _FormulaAwareCell:
    def __init__(self, value: object, data_type: str, coordinate: str) -> None:
        self.value = value
        self.data_type = data_type
        self.coordinate = coordinate


class _LazyWorkbook:
    worksheets = [_LazyWorksheet()]

    def close(self) -> None:
        pass


def test_lazy_workbook_iteration_error_is_wrapped_and_preview_is_cleaned_up(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setattr(excel_import_service, "load_workbook", lambda *args, **kwargs: _LazyWorkbook())
    temp_root = tmp_path / "imports"
    xlsx = tmp_path / "lazy.xlsx"
    xlsx.write_bytes(build_import_template())

    with pytest.raises(ValueError, match="Excel.*解析"):
        preview_import(xlsx, None, temp_root)

    assert list(temp_root.iterdir()) == []



def write_named_workbook(path: Path, content_rows: list[list[object]], test_rows: list[list[object]] | None = None) -> Path:
    workbook = Workbook()
    content = workbook.active
    content.title = "内容清单"
    content.append(list(excel_import_service.CONTENT_COLUMNS))
    for row in content_rows:
        content.append(row)
    if test_rows is not None:
        tests = workbook.create_sheet("测试场景")
        tests.append(list(excel_import_service.TEST_CASE_COLUMNS))
        for row in test_rows:
            tests.append(row)
    workbook.create_sheet("字段说明")
    workbook.save(path)
    return path


def tech_row(external_id: str, body: str = "普通正文") -> list[object]:
    return [external_id, "活动", "账号", "媒体", "小红书", "标题", body, None, None]


def test_trigger_words_are_not_required_by_text_only_precheck(tmp_path: Path) -> None:
    # Precheck only verifies the uploaded table can be read and has required
    # fields; semantic claims are handled by later text review.
    xlsx = write_named_workbook(
        tmp_path / "per-content.xlsx",
        [tech_row("triggered", "这是亲测内容"), tech_row("other")],
        [["other", "case-1", "通过", "指令", "结果", None, None, None, None, None, None, "other.png"]],
    )

    preview = preview_import(xlsx, None, tmp_path / "previews")

    assert preview.rows[0].valid is True
    assert not any("证据" in error or "测试场景" in error for error in preview.rows[0].errors)
    assert preview.rows[1].valid is True


def test_test_sheet_is_ignored_in_text_only_precheck(tmp_path: Path) -> None:
    xlsx = write_named_workbook(
        tmp_path / "orphan.xlsx",
        [tech_row("content-1")],
        [["missing", "case-1", "通过", "指令", "结果", None, None, None, None, None, None, "proof.png"]],
    )

    preview = preview_import(xlsx, None, tmp_path / "previews")

    assert preview.error_count == 0
    assert preview.errors == []
    assert preview.rows[0].valid is True
    assert preview.test_cases == []


def test_preview_requires_named_content_sheet(tmp_path: Path) -> None:
    xlsx = tmp_path / "legacy.xlsx"
    workbook = Workbook()
    workbook.active.append(list(IMPORT_COLUMNS))
    workbook.active.append(valid_row())
    workbook.save(xlsx)
    with pytest.raises(ValueError, match="内容清单"):
        preview_import(xlsx, None, tmp_path / "previews")


def test_malformed_xlsx_and_arbitrary_zip_have_stable_errors(tmp_path: Path) -> None:
    malformed = tmp_path / "malformed.xlsx"
    malformed.write_bytes(b"not an xlsx")
    with pytest.raises(ValueError, match="Excel|XLSX"):
        preview_import(malformed, None, tmp_path / "malformed-previews")

    arbitrary = tmp_path / "arbitrary.xlsx"
    write_zip(arbitrary, [("random.txt", b"not OOXML")])
    with pytest.raises(ValueError, match="Excel|XLSX"):
        preview_import(arbitrary, None, tmp_path / "arbitrary-previews")


def test_preview_identity_round_trips_and_is_strict(tmp_path: Path) -> None:
    identity_type = getattr(excel_import_service, "PreviewIdentity", None)
    assert identity_type is not None, "PreviewIdentity is missing"
    identity = identity_type(
        project_id=1, project_code="tech", content_type="TECH_MEDIA_REVIEW",
        package_version="0.9", supplier_id="supplier", batch_name="batch",
        project_type="科技媒体测评", owner_name="supplier",
    )
    xlsx = write_named_workbook(tmp_path / "identity.xlsx", [tech_row("content-1")], [])
    preview = preview_import(xlsx, None, tmp_path / "previews", identity=identity)
    loaded = load_preview(preview.token)
    assert loaded.identity == identity


def test_xlsx_uncompressed_expansion_limit_is_enforced(tmp_path: Path, monkeypatch) -> None:
    xlsx = tmp_path / "large.xlsx"
    xlsx.write_bytes(build_import_template())
    monkeypatch.setattr(excel_import_service, "MAX_XLSX_UNCOMPRESSED_BYTES", 1)
    with pytest.raises(ValueError, match="XLSX.*安全限制"):
        preview_import(xlsx, None, tmp_path / "previews")

from __future__ import annotations

import importlib
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from server.db import Base, create_db_engine
from server.models import (
    AgentResult,
    Asset,
    AuditRun,
    Batch,
    ContentItem,
    ContentVersion,
    FormatStatus,
    HumanDecision,
    Issue,
    PublishStatus,
    ReviewStatus,
    ReviewTask,
)
from server.seed import seed_default_project
from server.services import excel_import_service
from server.services.content_service import submit_batch
from server.services.excel_export_service import EXPORT_COLUMNS
from server.services.excel_import_service import CONTENT_COLUMNS, IMPORT_COLUMNS, preview_import
from server.services.review_service import run_audit
from scripts.text_review.reviewers.base import AgentReviewResult
from scripts.text_review.reviewers.tech_media import AGENT_ORDER, TechMediaReviewer


NEW_CONTENT_COLUMNS = ("标题", "内容", "类型", "目标平台", "作者", "发布日期", "图片/视频")


def confirm_import(*args, **kwargs):
    service = getattr(excel_import_service, "confirm_import", None)
    assert callable(service), "confirm_import service API is missing"
    return service(*args, **kwargs)


def export_batch(*args, **kwargs) -> bytes:
    try:
        module = importlib.import_module("server.services.excel_export_service")
    except ModuleNotFoundError as exc:
        pytest.fail(f"excel_export_service module is missing: {exc}")
    service = getattr(module, "export_batch", None)
    assert callable(service), "export_batch service API is missing"
    return service(*args, **kwargs)


@pytest.fixture(autouse=True)
def isolate_import_and_upload_roots(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("CONTENT_REVIEW_PREVIEW_ROOT_REGISTRY", str(tmp_path / "preview-roots.json"))
    monkeypatch.setenv("CR_DATA_DIR", str(tmp_path / "data"))
    excel_import_service._preview_locations.clear()


def make_session(tmp_path: Path) -> Session:
    engine = create_db_engine(f"sqlite:///{tmp_path / 'workflow.db'}")
    Base.metadata.create_all(engine)
    return Session(engine)


def write_workbook(path: Path, rows: list[list[object]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "内容清单"
    worksheet.append(list(IMPORT_COLUMNS))
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    return path


def supplier_row(
    external_id: object,
    *,
    campaign_theme: object = "暑期活动",
    platform: object = "小红书",
    title: object = "原始标题",
    body: object = "原始正文",
    image_filename: object = None,
    publish_time: object = "2026-07-20",
    note: object = "供应商备注",
) -> list[object]:
    return [external_id, campaign_theme, platform, title, body, image_filename, publish_time, note]


def write_content_workbook(path: Path, rows: list[list[object]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "内容清单"
    worksheet.append(list(CONTENT_COLUMNS))
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    return path


def content_row(
    external_id: object,
    *,
    campaign_theme: object = "暑期活动",
    account_name: object = "供应商账号",
    account_type: object = "媒体",
    platform: object = "小红书",
    title: object = "原始标题",
    body: object = "原始正文",
    image_filename: object = None,
    publish_time: object = "2026-07-20",
    note: object = "供应商备注",
) -> list[object]:
    return [external_id, campaign_theme, account_name, account_type, platform, title, body, image_filename, publish_time, note]


def write_zip(path: Path, entries: list[tuple[str, bytes]]) -> Path:
    with ZipFile(path, "w") as archive:
        for name, content in entries:
            archive.writestr(name, content, compress_type=ZIP_DEFLATED)
    return path


def payload_by_supplier_id(batch: Batch) -> dict[str, dict]:
    return {
        item.versions[0].payload["supplier_external_id"]: item.versions[0].payload
        for item in batch.content_items
    }


def test_new_format_confirm_persists_payload_media_and_exports_new_headers(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "内容清单"
    worksheet.append(list(NEW_CONTENT_COLUMNS))
    worksheet.append(["新标题", "新内容", "视频", "抖音", "地图作者", "2026-08-01", "cover.png"])
    xlsx = tmp_path / "new-format.xlsx"
    workbook.save(xlsx)
    archive = write_zip(tmp_path / "new-format.zip", [("cover.png", b"image")])
    preview = preview_import(xlsx, archive, tmp_path / "previews")

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(session, preview.token, project.id, "supplier", "新模板")
        item = batch.content_items[0]
        payload = item.versions[0].payload

        assert item.external_id == preview.rows[0].normalized["external_id"]
        assert payload["supplier_external_id"] == item.external_id
        assert payload["account_type"] == "视频"
        assert payload["platform"] == "抖音"
        assert payload["account_name"] == "地图作者"
        assert payload["publish_time"] == "2026-08-01"
        assert payload["image_filename"] == "cover.png"
        assert (tmp_path / "data" / "uploads" / payload["media"]).read_bytes() == b"image"

        exported = load_workbook(BytesIO(export_batch(session, batch.id))).active
        headers = [cell.value for cell in exported[1]]
        values = [cell.value for cell in exported[2]]
        assert headers[: len(NEW_CONTENT_COLUMNS)] == list(NEW_CONTENT_COLUMNS)
        assert values[: len(NEW_CONTENT_COLUMNS)] == [
            "新标题", "新内容", "视频", "抖音", "地图作者", "2026-08-01", "cover.png",
        ]


class _VisionTransport:
    def __init__(self, payload: dict):
        self.payload = payload

    def chat_json_multimodal(self, prompt: str, image_data_uri: str, schema: object) -> str:
        import json
        return json.dumps(self.payload, ensure_ascii=False)


class _CapturingPassReviewer(TechMediaReviewer):
    def __init__(self):
        super().__init__(llm=None)
        self.context = None

    def review_structured(self, context, profile):
        self.context = context
        return [
            AgentReviewResult.model_validate({
                "agent_id": agent_id,
                "agent_version": "tech-media-v1",
                "decision": "PASS",
                "summary": "pass",
                "score": 90,
                "confidence": 0.99,
                "issues": [],
            })
            for agent_id in AGENT_ORDER
        ]


def _new_image_workbook(path: Path, *, title: str, body: str, filename: str = "scene.png") -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "内容清单"
    sheet.append(list(NEW_CONTENT_COLUMNS))
    sheet.append([title, body, "图文", "小红书", "作者", "2026-08-01", filename])
    workbook.save(path)
    return path


def test_new_format_image_analysis_persists_and_flows_into_audit_context_and_issue(tmp_path: Path) -> None:
    png = b"\x89PNG\r\n\x1a\n" + b"test-scene"
    xlsx = _new_image_workbook(tmp_path / "image-analysis.xlsx", title="路线规划", body="路线规划说明")
    archive = write_zip(tmp_path / "image-analysis.zip", [("scene.png", png)])
    preview = preview_import(xlsx, archive, tmp_path / "previews")
    vision = _VisionTransport({
        "asset_id": "untrusted-model-id", "status": "ANALYZED", "is_test_scene": True,
        "visible_input": "北京南站到故宫", "visible_result": "三条路线", "visible_product": "百度地图",
        "detected_text": "北京南站 故宫", "confidence": 0.96,
        "missing_context": ["app_version", "tested_at"], "reasoning": "可见输入与路线结果",
    })

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(session, preview.token, project.id, "supplier", "image", image_llm=vision)
        item = batch.content_items[0]
        asset = session.scalars(select(Asset).where(Asset.content_item_id == item.id)).one()
        analysis = asset.asset_metadata["image_evidence_analysis"]
        assert analysis["asset_id"] == asset.asset_id
        assert analysis["status"] == "ANALYZED"
        assert analysis["is_test_scene"] is True
        assert analysis["verified"] is False

        reviewer = _CapturingPassReviewer()
        audit = run_audit(session, item.id, reviewer=reviewer)
        assert reviewer.context.image_evidence_analyses == [analysis]
        assert {issue.rule_id for issue in audit.issues} == {"IMAGE-EVIDENCE-CONTEXT-INCOMPLETE"}
        assert item.publish_status is PublishStatus.NOT_READY


def test_new_format_ordinary_screenshot_without_text_trigger_does_not_require_test_evidence(tmp_path: Path) -> None:
    png = b"\x89PNG\r\n\x1a\n" + b"ordinary-cover"
    xlsx = _new_image_workbook(tmp_path / "ordinary.xlsx", title="产品介绍", body="介绍地图产品功能")
    archive = write_zip(tmp_path / "ordinary.zip", [("scene.png", png)])
    preview = preview_import(xlsx, archive, tmp_path / "previews")
    vision = _VisionTransport({
        "asset_id": "ignored", "status": "ANALYZED", "is_test_scene": False,
        "visible_input": None, "visible_result": None, "visible_product": "百度地图",
        "detected_text": "百度地图", "confidence": 0.99, "missing_context": [],
        "reasoning": "普通封面，没有可见测试输入或结果",
    })

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(session, preview.token, project.id, "supplier", "ordinary", image_llm=vision)
        item = batch.content_items[0]
        reviewer = _CapturingPassReviewer()
        audit = run_audit(session, item.id, reviewer=reviewer)
        assert reviewer.context.image_evidence_analyses[0]["is_test_scene"] is False
        assert not any(issue.rule_id.startswith("IMAGE-EVIDENCE") or issue.rule_id == "TEST-EVIDENCE-001" for issue in audit.issues)
        assert item.publish_status is PublishStatus.READY


def test_confirm_import_retains_rows_missing_id_or_other_required_field(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "incomplete.xlsx",
        [
            supplier_row(" ", title="缺少编号"),
            supplier_row("missing-theme", campaign_theme=" ", title="缺少活动"),
        ],
    )
    preview = preview_import(xlsx, None, tmp_path / "previews")

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(
            session,
            preview.token,
            project_id=project.id,
            supplier_id="supplier-a",
            batch_name="保留不完整行",
        )

        assert len(batch.content_items) == 2
        by_title = {item.versions[0].title: item for item in batch.content_items}
        blank_id = by_title["缺少编号"]
        missing_theme = by_title["缺少活动"]
        assert blank_id.external_id == f"import:{preview.token[:16]}:row:2"
        assert blank_id.versions[0].payload["supplier_external_id"] is None
        assert blank_id.format_status == FormatStatus.INCOMPLETE
        assert missing_theme.external_id == f"import:{preview.token[:16]}:row:3"
        assert missing_theme.versions[0].payload["supplier_external_id"] == "missing-theme"
        assert missing_theme.versions[0].payload["campaign_theme"] is None
        assert missing_theme.format_status == FormatStatus.INCOMPLETE


def test_confirm_import_imports_multiple_preview_rows_into_one_batch(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "multiple.xlsx",
        [supplier_row("content-1", title="标题一"), supplier_row("content-2", title="标题二")],
    )
    preview = preview_import(xlsx, None, tmp_path / "previews")
    assert [(row.manuscript_index, row.row_number) for row in preview.rows] == [(1, 2), (2, 3)]

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(
            session,
            preview.token,
            project_id=project.id,
            supplier_id="supplier-a",
            batch_name="Excel 批量导入",
        )

        assert batch.import_token == preview.token
        assert batch.project_id == project.id
        assert batch.supplier_id == "supplier-a"
        assert batch.name == "Excel 批量导入"
        assert len(batch.content_items) == 2
        assert {item.external_id for item in batch.content_items} == {"content-1", "content-2"}
        assert {item.format_status for item in batch.content_items} == {FormatStatus.PASSED}
        assert all(item.review_status is ReviewStatus.NOT_STARTED for item in batch.content_items)
        assert all(item.publish_status is PublishStatus.NOT_READY for item in batch.content_items)
        assert all(len(item.versions) == 1 for item in batch.content_items)
        assert all(item.versions[0].source == "SUPPLIER" for item in batch.content_items)
        payloads = payload_by_supplier_id(batch)
        assert payloads["content-1"]["row_number"] == 2
        assert payloads["content-2"]["row_number"] == 3
        assert payloads["content-1"]["campaign_theme"] == "暑期活动"
        assert not (tmp_path / "previews" / preview.token).exists()


def test_confirm_import_retains_preview_error_rows_with_errors_in_payload(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "duplicates.xlsx",
        [supplier_row("duplicate", title="第一条"), supplier_row("duplicate", title="第二条")],
    )
    preview = preview_import(xlsx, None, tmp_path / "previews")
    assert preview.error_count == 2

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(
            session,
            preview.token,
            project_id=project.id,
            supplier_id="supplier-a",
            batch_name="包含错误行",
        )

        assert len(batch.content_items) == 2
        assert all(item.format_status is not FormatStatus.PASSED for item in batch.content_items)
        assert {item.format_status for item in batch.content_items} == {FormatStatus.INVALID}
        assert "duplicate" not in {item.external_id for item in batch.content_items}
        assert len({item.external_id for item in batch.content_items}) == 2
        for item in batch.content_items:
            payload = item.versions[0].payload
            assert payload["supplier_external_id"] == "duplicate"
            assert any("重复" in error for error in payload["preview_errors"])
            assert payload["preview_warnings"] == []


def test_confirm_import_copies_referenced_preview_images_by_exact_filename_per_row(tmp_path: Path) -> None:
    xlsx = write_workbook(
        tmp_path / "images.xlsx",
        [
            supplier_row("row-second", image_filename="second.png"),
            supplier_row("row-first", image_filename="first.png"),
        ],
    )
    archive = write_zip(tmp_path / "images.zip", [("first.png", b"first-image"), ("second.png", b"second-image")])
    preview = preview_import(xlsx, archive, tmp_path / "previews")
    assert preview.valid_count == 2

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(
            session,
            preview.token,
            project_id=project.id,
            supplier_id="supplier-images",
            batch_name="图片导入",
        )

        payloads = payload_by_supplier_id(batch)
        uploads_dir = tmp_path / "data" / "uploads"
        assert payloads["row-second"]["image_filename"] == "second.png"
        assert payloads["row-first"]["image_filename"] == "first.png"
        assert (uploads_dir / payloads["row-second"]["media"]).read_bytes() == b"second-image"
        assert (uploads_dir / payloads["row-first"]["media"]).read_bytes() == b"first-image"
        assert Path(payloads["row-second"]["media"]).name == payloads["row-second"]["media"]
        assert Path(payloads["row-first"]["media"]).name == payloads["row-first"]["media"]


def test_confirm_import_keeps_committed_media_when_refresh_fails_after_commit(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    xlsx = write_workbook(
        tmp_path / "refresh-failure.xlsx",
        [supplier_row("row-with-image", image_filename="photo.png")],
    )
    archive = write_zip(tmp_path / "refresh-failure.zip", [("photo.png", b"durable-image")])
    preview = preview_import(xlsx, archive, tmp_path / "previews")

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        original_refresh = session.refresh
        failed = False

        def fail_first_batch_refresh(instance, *args, **kwargs):
            nonlocal failed
            if isinstance(instance, Batch) and not failed:
                failed = True
                raise RuntimeError("connection lost after commit")
            return original_refresh(instance, *args, **kwargs)

        monkeypatch.setattr(session, "refresh", fail_first_batch_refresh)

        with pytest.raises(RuntimeError, match="connection lost after commit"):
            confirm_import(
                session,
                preview.token,
                project_id=project.id,
                supplier_id="supplier-images",
                batch_name="图片导入",
            )

        committed = session.scalar(select(Batch).where(Batch.import_token == preview.token))
        assert committed is not None
        payload = payload_by_supplier_id(committed)["row-with-image"]
        media_path = tmp_path / "data" / "uploads" / payload["media"]
        assert media_path.read_bytes() == b"durable-image"
        assert (tmp_path / "previews" / preview.token).exists()

        retry = confirm_import(
            session,
            preview.token,
            project_id=project.id,
            supplier_id="supplier-images",
            batch_name="图片导入",
        )

        assert retry.id == committed.id
        assert session.scalar(select(func.count(Batch.id))) == 1
        assert media_path.read_bytes() == b"durable-image"


def test_confirm_import_is_idempotent_for_same_token_without_duplicate_content(tmp_path: Path) -> None:
    xlsx = write_workbook(tmp_path / "idempotent.xlsx", [supplier_row("content-1")])
    preview = preview_import(xlsx, None, tmp_path / "previews")

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        first = confirm_import(
            session,
            preview.token,
            project_id=project.id,
            supplier_id="supplier-a",
            batch_name="首次确认",
        )
        first_id = first.id
        first_content_ids = [item.id for item in first.content_items]

        second = confirm_import(
            session,
            preview.token,
            project_id=project.id,
            supplier_id="supplier-a",
            batch_name="首次确认",
        )

        assert second.id == first_id
        assert [item.id for item in second.content_items] == first_content_ids
        assert session.scalar(select(func.count(Batch.id))) == 1
        assert session.scalar(select(func.count(ContentItem.id))) == 1


def test_confirm_import_returns_existing_batch_after_concurrent_import_token_conflict(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    engine = create_db_engine(f"sqlite:///{tmp_path / 'workflow.db'}")
    Base.metadata.create_all(engine)
    with Session(engine) as setup_session:
        project = seed_default_project(setup_session)
        project_id = project.id
        setup_session.commit()

    xlsx = write_workbook(
        tmp_path / "integrity-conflict.xlsx",
        [supplier_row("losing-row", image_filename="photo.png")],
    )
    archive = write_zip(tmp_path / "integrity-conflict.zip", [("photo.png", b"losing-image")])
    preview = preview_import(xlsx, archive, tmp_path / "previews")
    winner_id: int | None = None

    with Session(engine) as session:
        original_flush = session.flush
        inserted_winner = False

        def insert_winner_then_flush(*args, **kwargs):
            nonlocal inserted_winner, winner_id
            has_losing_batch = any(
                isinstance(instance, Batch) and instance.import_token == preview.token
                for instance in session.new
            )
            if has_losing_batch and not inserted_winner:
                inserted_winner = True
                with Session(engine) as winner_session:
                    winner = submit_batch(
                        winner_session,
                        project_id=project_id,
                        supplier_id="winner",
                        name="已提交批次",
                        contents=[
                            {
                                "external_id": "winner-row",
                                "title": "已提交标题",
                                "body": "已提交正文",
                                "payload": {"supplier_external_id": "winner-row"},
                            }
                        ],
                        import_token=preview.token,
                    )
                    winner_id = winner.id
            return original_flush(*args, **kwargs)

        monkeypatch.setattr(session, "flush", insert_winner_then_flush)

        batch = confirm_import(
            session,
            preview.token,
            project_id=project_id,
            supplier_id="loser",
            batch_name="并发失败批次",
        )

        assert winner_id is not None
        assert batch.id == winner_id
        assert batch.supplier_id == "winner"
        assert session.scalar(select(func.count(Batch.id))) == 1
        assert session.scalar(select(func.count(ContentItem.id))) == 1
        uploads_dir = tmp_path / "data" / "uploads"
        assert not uploads_dir.exists() or list(uploads_dir.iterdir()) == []


def test_export_batch_returns_xlsx_with_supplier_and_review_columns(tmp_path: Path) -> None:
    completed_at = datetime(2026, 7, 15, 8, 30, 0)
    xlsx = write_content_workbook(
        tmp_path / "export.xlsx",
        [
            content_row(
                "supplier-1",
                campaign_theme="新品活动",
                account_name="地图测评号",
                account_type="科技媒体",
                platform="抖音",
                title="供应商标题",
                body="供应商正文",
                image_filename=None,
                publish_time="2026-08-01",
                note="导出备注",
            )
        ],
    )
    preview = preview_import(xlsx, None, tmp_path / "previews")

    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(
            session,
            preview.token,
            project_id=project.id,
            supplier_id="supplier-a",
            batch_name="导出批次",
        )
        item = batch.content_items[0]
        supplier_version = item.versions[0]
        audit = AuditRun(
            content_item=item,
            content_version=supplier_version,
            rule_version=project.current_rule_version,
            model="model-x",
            prompt_version="prompt-x",
            status="COMPLETED",
            completed_at=completed_at,
        )
        result = AgentResult(
            audit_run=audit,
            agent_name="quality",
            status="COMPLETED",
            raw_result={"source": "test"},
        )
        low_issue = Issue(
            audit_run=audit,
            agent_result=result,
            rule_id="QUALITY-1",
            category="quality",
            severity="low",
            field="body",
            evidence_quote="证据一",
            reason="原因一",
            suggestion="建议一",
            auto_fixable=True,
            human_required=False,
            confidence=0.9,
        )
        high_issue = Issue(
            audit_run=audit,
            agent_result=result,
            rule_id="RISK-1",
            category="external",
            severity="high",
            field="title",
            evidence_quote="证据二",
            reason="原因二",
            suggestion="建议二",
            auto_fixable=False,
            human_required=True,
            confidence=0.99,
        )
        task = ReviewTask(
            content_item=item,
            issue=high_issue,
            target_content_version=supplier_version,
            audit_run=audit,
            task_type="RISK_REVIEW",
            status="CLOSED",
            closed_at=completed_at,
        )
        HumanDecision(
            review_task=task,
            decision="APPROVE_RISK",
            reviewer="legal@example.com",
            note="已确认",
            payload={},
        )
        final_version = ContentVersion(
            content_item=item,
            version=2,
            source="HUMAN_APPROVED",
            title="最终标题",
            body="最终正文",
            payload=supplier_version.payload,
        )
        item.title = "最终标题"
        item.review_status = ReviewStatus.PASSED
        item.publish_status = PublishStatus.READY
        session.add_all([audit, low_issue, final_version])
        session.commit()

        workbook = load_workbook(BytesIO(export_batch(session, batch.id)))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        values = {headers[index]: worksheet.cell(row=2, column=index + 1).value for index in range(len(headers))}

        assert headers == list(NEW_CONTENT_COLUMNS) + list(EXPORT_COLUMNS)
        assert values["标题"] == "供应商标题"
        assert values["内容"] == "供应商正文"
        assert values["类型"] == "科技媒体"
        assert values["目标平台"] == "抖音"
        assert values["作者"] == "地图测评号"
        assert values["发布日期"] == "2026-08-01"
        assert values["图片/视频"] is None
        assert values["系统内容编号"] == item.id
        assert values["批次编号"] == batch.id
        assert values["格式校验"] == "PASSED"
        assert values["审核状态"] == "PASSED"
        assert values["发布状态"] == "READY"
        assert values["问题数量"] == 2
        assert values["最高风险等级"] == "HIGH"
        assert values["问题分类"] == "quality\nexternal"
        assert values["命中规则"] == "QUALITY-1\nRISK-1"
        assert values["问题原因"] == "原因一\n原因二"
        assert values["原文证据"] == "证据一\n证据二"
        assert values["修改建议"] == "建议一\n建议二"
        assert values["最终标题"] == "最终标题"
        assert values["最终正文"] == "最终正文"
        assert values["最终版本来源"] == "HUMAN_APPROVED"
        assert values["是否需要人工"] == "是"
        assert values["是否已人工确认"] == "是"
        assert values["审核模型"] == "model-x"
        assert values["规则版本"] == project.current_rule_version.version
        assert values["审核完成时间"] == completed_at
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == f"A1:{get_column_letter(len(headers))}2"
        assert worksheet["E2"].alignment.wrap_text is True
        assert worksheet["AA2"].alignment.wrap_text is True


def test_export_batch_sanitizes_formula_like_strings(tmp_path: Path) -> None:
    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = Batch(project=project, supplier_id="supplier-a", name="公式注入导出")
        item = ContentItem(
            project=project,
            batch=batch,
            external_id="safe-system-id",
            title="@供应商标题",
            format_status=FormatStatus.PASSED,
        )
        supplier_payload = {
            "supplier_external_id": "=supplier-id",
            "campaign_theme": "+campaign",
            "account_name": "=account-name",
            "account_type": "@account-type",
            "platform": "-platform",
            "title": "@title",
            "body": "\tbody",
            "image_filename": "\rimage.png",
            "publish_time": "=2026-08-01",
            "note": "+note",
        }
        supplier_version = ContentVersion(
            content_item=item,
            version=1,
            source="SUPPLIER",
            title="@title",
            body="\tbody",
            payload=supplier_payload,
        )
        audit = AuditRun(
            content_item=item,
            content_version=supplier_version,
            rule_version=project.current_rule_version,
            model="@model",
            prompt_version="prompt-x",
            status="COMPLETED",
        )
        result = AgentResult(
            audit_run=audit,
            agent_name="quality",
            status="COMPLETED",
            raw_result={},
        )
        Issue(
            audit_run=audit,
            agent_result=result,
            rule_id="+RULE-1",
            category="-category",
            severity="high",
            field="body",
            evidence_quote="\tevidence",
            reason="=reason",
            suggestion="\rsuggestion",
            auto_fixable=False,
            human_required=True,
            confidence=0.99,
        )
        ContentVersion(
            content_item=item,
            version=2,
            source="=HUMAN_APPROVED",
            title="+final title",
            body="-final body",
            payload=supplier_payload,
        )
        session.add(batch)
        session.commit()

        workbook = load_workbook(BytesIO(export_batch(session, batch.id)), data_only=False)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        expected_literals = {
            "标题": "'@title",
            "内容": "'\tbody",
            "类型": "'@account-type",
            "目标平台": "'-platform",
            "作者": "'=account-name",
            "发布日期": "'=2026-08-01",
            "图片/视频": "'\rimage.png",
            "问题分类": "'-category",
            "命中规则": "'+RULE-1",
            "问题原因": "'=reason",
            "原文证据": "'\tevidence",
            "修改建议": "'\rsuggestion",
            "最终标题": "'+final title",
            "最终正文": "'-final body",
            "最终版本来源": "'=HUMAN_APPROVED",
            "审核模型": "'@model",
        }
        for header, expected in expected_literals.items():
            cell = worksheet.cell(row=2, column=headers.index(header) + 1)
            assert cell.data_type == "s", header
            assert cell.value == expected


def test_export_batch_raises_for_missing_batch(tmp_path: Path) -> None:
    with make_session(tmp_path) as session:
        with pytest.raises(ValueError, match="Batch 999 does not exist"):
            export_batch(session, 999)



def _named_shared_evidence_workbook(path: Path, filenames: list[str]) -> Path:
    workbook = Workbook()
    content = workbook.active
    content.title = "内容清单"
    content.append(list(excel_import_service.CONTENT_COLUMNS))
    content.append(["content-1", "活动", "账号", "媒体", "小红书", "标题", "普通正文", None, None, None])
    tests = workbook.create_sheet("测试场景")
    tests.append(list(excel_import_service.TEST_CASE_COLUMNS))
    for index, filename in enumerate(filenames, 1):
        tests.append(["content-1", f"case-{index}", "通过", "指令", "结果", None, None, None, None, None, None, filename])
    workbook.create_sheet("字段说明")
    workbook.save(path)
    return path


def test_confirm_reuses_shared_evidence_asset_and_copies_once(tmp_path: Path) -> None:
    xlsx = _named_shared_evidence_workbook(tmp_path / "shared.xlsx", ["proof.png", "proof.png"])
    archive = write_zip(tmp_path / "shared.zip", [("proof.png", b"shared")])
    preview = preview_import(xlsx, archive, tmp_path / "previews")
    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(session, preview.token, project.id, "supplier", "batch")
        item = batch.content_items[0]
        assert len(item.assets) == 1
        assert len(item.test_cases) == 2
        assert {binding.asset_id for test in item.test_cases for binding in test.evidence} == {item.assets[0].id}
        assert len(list((tmp_path / "data" / "uploads").iterdir())) == 1


@pytest.mark.parametrize(
    ("filename", "kind", "mime"),
    [("proof.gif", "SCREENSHOT", "image/gif"), ("proof.webm", "SCREEN_RECORDING", "video/webm"), ("proof.json", "TEST_LOG", "application/json")],
)
def test_confirm_maps_evidence_asset_kind_and_mime(tmp_path: Path, filename: str, kind: str, mime: str) -> None:
    xlsx = _named_shared_evidence_workbook(tmp_path / f"{kind}.xlsx", [filename])
    archive = write_zip(tmp_path / f"{kind}.zip", [(filename, b"evidence")])
    preview = preview_import(xlsx, archive, tmp_path / f"previews-{kind}")
    with make_session(tmp_path) as session:
        project = seed_default_project(session)
        batch = confirm_import(session, preview.token, project.id, "supplier", kind)
        asset = batch.content_items[0].assets[0]
        assert asset.kind.value == kind
        assert asset.mime_type == mime


def test_shared_severity_helper_ranks_critical_and_unknown_safely() -> None:
    module = importlib.import_module("server.services.severity_service")
    assert module.highest_severity(["LOW", "CRITICAL"]) == "CRITICAL"
    assert module.highest_severity(["MEDIUM", "HIGH"]) == "HIGH"
    assert module.severity_rank("UNKNOWN") == module.severity_rank("HIGH")

from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from xml.sax.saxutils import escape
from zipfile import ZipFile

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from server import main
from server.db import create_db_engine
from server.models import AgentResult, Asset, AuditRun, ContentItem, FormatStatus, Issue, Project
from server.services import excel_import_service
from server.services.excel_export_service import EXPORT_COLUMNS
from server.services.excel_import_service import CONTENT_COLUMNS, TEST_CASE_COLUMNS


NEW_CONTENT_COLUMNS = ("标题", "内容", "类型", "目标平台", "作者", "发布日期", "图片/视频")


@pytest.fixture
def excel_api(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    database_url = f"sqlite:///{tmp_path / 'excel-api.db'}"
    monkeypatch.setenv("DATABASE_URL", database_url)
    monkeypatch.setenv("CR_DATA_DIR", str(tmp_path / "data"))
    monkeypatch.setenv("CONTENT_REVIEW_PREVIEW_ROOT_REGISTRY", str(tmp_path / "preview-roots.json"))
    monkeypatch.setenv("INITIAL_ADMIN_USERNAME", "test-admin")
    monkeypatch.setenv("INITIAL_ADMIN_PASSWORD", "test-admin-password")
    monkeypatch.setenv("SESSION_SECRET", "test-session-secret-with-at-least-32-bytes")
    monkeypatch.setenv("TRUSTED_PUBLIC_ORIGINS", "http://testserver")
    engine = create_db_engine(database_url)

    def test_session():
        with Session(engine) as session:
            yield session

    main.app.dependency_overrides[main.get_session] = test_session
    with TestClient(main.app) as client:
        authenticated = client.post(
            "/api/auth/login",
            json={"username": "test-admin", "password": "test-admin-password"},
        )
        assert authenticated.status_code == 200
        client.headers.update({
            "Origin": "http://testserver",
            "X-CSRF-Token": authenticated.json()["csrf_token"],
        })
        yield client, engine, tmp_path
    main.app.dependency_overrides.clear()
    excel_import_service._preview_locations.clear()
    engine.dispose()


def workbook_bytes(
    *,
    external_id: str = "content-1",
    campaign_theme: str = "活动",
    account_name: str = "账号",
    account_type: str = "媒体",
    body: str = "普通正文",
    evidence: str | None = None,
) -> bytes:
    workbook = Workbook()
    content = workbook.active
    content.title = "内容清单"
    content.append(list(CONTENT_COLUMNS))
    content.append([external_id, campaign_theme, account_name, account_type, "小红书", "原始标题", body, None, "2026-07-20", "备注"])
    tests = workbook.create_sheet("测试场景")
    tests.append(list(TEST_CASE_COLUMNS))
    if evidence:
        tests.append([external_id, "case-1", "通过", "打开地图", "返回结果", "北京", None, "1.0", "设备", "系统", "WiFi", evidence])
    workbook.create_sheet("字段说明")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def new_workbook_bytes() -> bytes:
    workbook = Workbook()
    content = workbook.active
    content.title = "内容清单"
    content.append(list(NEW_CONTENT_COLUMNS))
    content.append(["新标题", "新内容", "图文", "小红书", "新作者", "2026-07-21", None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def zip_bytes(filename: str, content: bytes = b"evidence") -> bytes:
    output = BytesIO()
    with ZipFile(output, "w") as archive:
        archive.writestr(filename, content)
    return output.getvalue()


def docx_bytes(text: str) -> bytes:
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        f"<w:p><w:r><w:t>{escape(text)}</w:t></w:r></w:p>"
        "</w:body>"
        "</w:document>"
    )
    output = BytesIO()
    with ZipFile(output, "w") as archive:
        archive.writestr("word/document.xml", document)
    return output.getvalue()


def project(client: TestClient) -> dict:
    return client.get("/api/projects").json()[0]


def preview_request(client: TestClient, project_id: int, *, xlsx: bytes | None = None, zip_content: bytes | None = None, supplier: str = "supplier", batch: str = "batch", brief: str = "本批次 Brief：只允许路线规划和沿途信息查询，不允许写自动订酒店。"):
    files = {"excel_file": ("input.xlsx", xlsx or workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    if zip_content is not None:
        files["evidence_zip"] = ("evidence.zip", zip_content, "application/zip")
    return client.post(
        "/api/imports/preview",
        data={
            "project_id": str(project_id),
            "supplier_id": supplier,
            "batch_name": batch,
            "review_brief": brief,
        },
        files=files,
    )


def test_template_route_has_attachment_and_exact_sheets(excel_api) -> None:
    client, _, _ = excel_api
    response = client.get("/api/import-template")
    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["内容清单", "测试场景", "字段说明"]
    assert tuple(cell.value for cell in workbook["内容清单"][1]) == NEW_CONTENT_COLUMNS
    assert tuple(cell.value for cell in workbook["测试场景"][1]) == TEST_CASE_COLUMNS


def test_preview_multipart_accepts_new_content_headers(excel_api) -> None:
    client, _, _ = excel_api
    current = project(client)

    response = preview_request(client, current["id"], xlsx=new_workbook_bytes())

    assert response.status_code == 200, response.text
    row = response.json()["rows"][0]
    assert row["valid"] is True
    assert row["normalized"]["title"] == "新标题"
    assert row["normalized"]["body"] == "新内容"
    assert row["normalized"]["external_id"].startswith("excel:")


def test_confirm_uses_configured_reviewer_llm_for_image_analysis(excel_api) -> None:
    client, engine, _ = excel_api
    current = project(client)
    workbook = Workbook()
    content = workbook.active
    content.title = "内容清单"
    content.append(list(NEW_CONTENT_COLUMNS))
    content.append(["路线规划", "普通产品介绍", "图文", "小红书", "作者", "2026-07-21", "scene.png"])
    output = BytesIO()
    workbook.save(output)

    class VisionTransport:
        def chat_json_multimodal(self, prompt: str, image_data_uri: str, schema: object) -> str:
            return json.dumps({
                "asset_id": "model-value-is-ignored",
                "status": "ANALYZED",
                "is_test_scene": False,
                "visible_input": None,
                "visible_result": None,
                "visible_product": "百度地图",
                "detected_text": "百度地图",
                "confidence": 0.98,
                "missing_context": [],
                "reasoning": "普通产品截图",
            }, ensure_ascii=False)

    main.app.dependency_overrides[main.get_audit_reviewer] = lambda: SimpleNamespace(llm=VisionTransport())
    preview = preview_request(
        client,
        current["id"],
        xlsx=output.getvalue(),
        zip_content=zip_bytes("scene.png", b"\x89PNG\r\n\x1a\nimage"),
    )
    token = preview.json()["token"]
    confirmed = client.post(
        f"/api/imports/{token}/confirm",
        json={"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"},
    )

    assert confirmed.status_code == 200, confirmed.text
    with Session(engine) as session:
        asset = session.scalar(select(Asset))
        assert asset is not None
        assert asset.asset_metadata["image_evidence_analysis"]["status"] == "ANALYZED"


def test_preview_multipart_returns_typed_identity_rows_and_tests(excel_api) -> None:
    client, _, _ = excel_api
    current = project(client)
    response = preview_request(client, current["id"], xlsx=workbook_bytes(body="亲测 1 个测试", evidence="proof.png"), zip_content=zip_bytes("proof.png"))
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["project_id"] == current["id"]
    assert payload["supplier_id"] == "supplier"
    assert payload["batch_name"] == "batch"
    assert payload["project_type"] == current["name"]
    assert payload["owner_name"] == "supplier"
    assert payload["rows"][0]["tests"][0]["external_test_case_id"] == "case-1"
    assert payload["errors"] == []


def test_preview_and_confirm_persist_batch_specific_brief(excel_api) -> None:
    client, _, _ = excel_api
    current = project(client)
    brief = "2026 春季口径：仅允许多点路线规划，不允许出现自动订酒店。"
    response = preview_request(client, current["id"], brief=brief)
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["review_brief"] == brief
    assert "自动订酒店" in payload["brief_summary"]

    confirmed = client.post(
        f"/api/imports/{payload['token']}/confirm",
        json={"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"},
    )
    assert confirmed.status_code == 200, confirmed.text
    assert confirmed.json()["review_brief"] == brief
    assert confirmed.json()["project_type"] == current["name"]
    assert confirmed.json()["owner_name"] == "supplier"


def test_preview_accepts_docx_brief_file_and_persists_identity(excel_api) -> None:
    client, _, _ = excel_api
    current = project(client)
    files = {
        "excel_file": ("input.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "brief_file": (
            "brief.docx",
            docx_bytes("Word Brief：本批次只允许写新路线功能，不允许写旧活动口径。"),
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ),
    }
    response = client.post(
        "/api/imports/preview",
        data={
            "project_id": str(current["id"]),
            "supplier_id": "owner-a",
            "batch_name": "2026-春季-01",
            "project_type": "新品功能稿",
            "owner_name": "owner-a",
        },
        files=files,
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert "新路线功能" in payload["review_brief"]
    assert payload["project_type"] == "新品功能稿"
    assert payload["owner_name"] == "owner-a"

    confirmed = client.post(
        f"/api/imports/{payload['token']}/confirm",
        json={
            "project_id": current["id"],
            "supplier_id": "owner-a",
            "batch_name": "2026-春季-01",
            "project_type": "新品功能稿",
            "owner_name": "owner-a",
        },
    )
    assert confirmed.status_code == 200, confirmed.text
    body = confirmed.json()
    assert body["project_type"] == "新品功能稿"
    assert body["owner_name"] == "owner-a"


def test_preview_rejects_non_tech_bad_suffix_malformed_and_oversize(excel_api, monkeypatch: pytest.MonkeyPatch) -> None:
    client, engine, _ = excel_api
    with Session(engine) as session:
        other = Project(name="Other", code="other", content_type="OTHER")
        session.add(other); session.commit(); other_id = other.id
    assert preview_request(client, other_id).status_code == 422
    current = project(client)
    bad_suffix = client.post("/api/imports/preview", data={"project_id": current["id"], "supplier_id": "s", "batch_name": "b", "review_brief": "批次 Brief"}, files={"excel_file": ("input.xls", b"x", "application/octet-stream")})
    assert bad_suffix.status_code == 422
    assert preview_request(client, current["id"], xlsx=b"not-xlsx").status_code == 422
    monkeypatch.setattr(main, "MAX_EXCEL_UPLOAD_BYTES", 4)
    assert preview_request(client, current["id"]).status_code == 413


def test_preview_temp_upload_directory_is_cleaned(excel_api) -> None:
    client, _, tmp_path = excel_api
    current = project(client)
    assert preview_request(client, current["id"]).status_code == 200
    root = tmp_path / "data" / "import-previews"
    assert not any(path.name.startswith("upload-") for path in root.iterdir())
    assert not any(path.suffix == ".xlsx" for path in root.rglob("*"))


def test_confirm_is_idempotent_and_rejects_cross_identity(excel_api) -> None:
    client, _, _ = excel_api
    current = project(client)
    token = preview_request(client, current["id"]).json()["token"]
    body = {"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"}
    first = client.post(f"/api/imports/{token}/confirm", json=body)
    second = client.post(f"/api/imports/{token}/confirm", json=body)
    assert first.status_code == second.status_code == 200
    assert first.json()["id"] == second.json()["id"]
    assert client.post(f"/api/imports/{token}/confirm", json={**body, "supplier_id": "other"}).status_code == 422
    assert client.post("/api/imports/not-a-token/confirm", json=body).status_code == 422


@pytest.mark.parametrize(
    ("xlsx", "missing_field", "supplier_external_id"),
    [
        (workbook_bytes(external_id=" "), "供应商内容编号", None),
        (workbook_bytes(external_id="missing-theme", campaign_theme=" "), "活动主题", "missing-theme"),
    ],
    ids=["missing-content-id", "missing-campaign-theme"],
)
def test_http_preview_and_confirm_retain_incomplete_rows(
    excel_api, xlsx: bytes, missing_field: str, supplier_external_id: str | None
) -> None:
    client, engine, _ = excel_api
    current = project(client)

    preview = preview_request(client, current["id"], xlsx=xlsx)
    assert preview.status_code == 200, preview.text
    payload = preview.json()
    assert payload["rows"][0]["valid"] is False
    assert any(missing_field in error for error in payload["rows"][0]["errors"])

    confirmed = client.post(
        f"/api/imports/{payload['token']}/confirm",
        json={"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"},
    )
    assert confirmed.status_code == 200, confirmed.text
    content_id = confirmed.json()["contents"][0]["id"]
    with Session(engine) as session:
        item = session.get(ContentItem, content_id)
        assert item is not None
        assert item.external_id == f"import:{payload['token'][:16]}:row:2"
        assert item.format_status == FormatStatus.INCOMPLETE
        assert item.versions[0].payload["supplier_external_id"] == supplier_external_id


def test_expired_token_returns_422(excel_api) -> None:
    client, _, tmp_path = excel_api
    current = project(client)
    token = preview_request(client, current["id"]).json()["token"]
    manifest = tmp_path / "data" / "import-previews" / token / "preview.json"
    payload = json.loads(manifest.read_text(encoding="utf-8"))
    payload["expires_at"] = (datetime.now(timezone.utc) - timedelta(seconds=1)).isoformat()
    manifest.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    response = client.post(f"/api/imports/{token}/confirm", json={"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"})
    assert response.status_code == 422


def test_export_route_and_missing_batch(excel_api) -> None:
    client, _, _ = excel_api
    current = project(client)
    token = preview_request(
        client,
        current["id"],
        xlsx=workbook_bytes(account_name="地图账号", account_type="科技媒体"),
    ).json()["token"]
    batch = client.post(f"/api/imports/{token}/confirm", json={"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"}).json()
    exported = client.get(f"/api/batches/{batch['id']}/export")
    assert exported.status_code == 200
    assert "attachment" in exported.headers["content-disposition"]
    sheet = load_workbook(BytesIO(exported.content)).active
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]
    assert headers == list(NEW_CONTENT_COLUMNS) + list(EXPORT_COLUMNS)
    assert values[: len(NEW_CONTENT_COLUMNS)] == [
        "原始标题",
        "普通正文",
        "科技媒体",
        "小红书",
        "地图账号",
        "2026-07-20",
        None,
    ]
    assert client.get("/api/batches/99999/export").status_code == 404


def test_contents_table_contract_filters_severity_agents_and_missing_media(excel_api) -> None:
    client, engine, _ = excel_api
    current = project(client)
    token = preview_request(client, current["id"]).json()["token"]
    batch = client.post(f"/api/imports/{token}/confirm", json={"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"}).json()
    content_id = batch["contents"][0]["id"]
    with Session(engine) as session:
        item = session.get(main.ContentItem, content_id)
        audit = AuditRun(content_item=item, content_version=item.versions[0], rule_version=item.project.current_rule_version, model="m", prompt_version="p", status="COMPLETED")
        for agent_id in reversed(["CONTENT_QUALITY", "COMPLIANCE", "BRAND", "PRODUCT_ACCURACY", "CAMPAIGN_EFFECTIVENESS"]):
            audit.agent_results.append(AgentResult(agent_name=agent_id, agent_id=agent_id, agent_version="v1", decision="PASS", summary=agent_id, score=90, status="COMPLETED", raw_result={}))
        audit.issues.extend([
            Issue(rule_id="LOW", category="quality", severity="LOW", field="body", evidence_quote="", reason="low", suggestion="s1", auto_fixable=False, human_required=False, confidence=1),
            Issue(rule_id="CRITICAL", category="risk", severity="CRITICAL", field="body", evidence_quote="", reason="critical", suggestion="s2", auto_fixable=False, human_required=True, confidence=1),
        ])
        session.add(audit); session.commit()
    response = client.get("/api/contents/table", params={"project_id": current["id"], "batch_id": batch["id"], "format_status": "PASSED"})
    assert response.status_code == 200
    row = response.json()[0]
    assert row["supplier_external_id"] == "content-1"
    assert row["original_title"] == "原始标题" and row["final_title"] == "原始标题"
    assert row["highest_severity"] == "CRITICAL"
    assert [agent["agent_id"] for agent in row["agents"]] == ["CONTENT_QUALITY", "COMPLIANCE", "BRAND", "PRODUCT_ACCURACY", "CAMPAIGN_EFFECTIVENESS"]
    assert row["media_url"] is None
    assert client.get("/api/contents/table", params={"review_status": "INVALID"}).status_code == 422


def test_contents_table_always_returns_active_canonical_agent_slots(excel_api) -> None:
    client, _, _ = excel_api
    current = project(client)
    token = preview_request(client, current["id"]).json()["token"]
    client.post(f"/api/imports/{token}/confirm", json={"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"})
    row = client.get("/api/contents/table").json()[0]
    assert [agent["agent_id"] for agent in row["agents"]] == ["CONTENT_QUALITY", "COMPLIANCE", "BRAND", "PRODUCT_ACCURACY", "CAMPAIGN_EFFECTIVENESS"]
    assert all(agent["status"] == "NOT_RUN" for agent in row["agents"])


def test_preview_orphan_test_does_not_block_confirmation(excel_api) -> None:
    client, _, _ = excel_api
    current = project(client)
    data = workbook_bytes()
    workbook = load_workbook(BytesIO(data))
    tests = workbook["测试场景"]
    tests.append(["missing", "orphan", "通过", "指令", "结果", None, None, None, None, None, None, "proof.png"])
    output = BytesIO(); workbook.save(output)
    response = preview_request(client, current["id"], xlsx=output.getvalue(), zip_content=zip_bytes("proof.png"))
    assert response.status_code == 200
    payload = response.json()
    assert payload["errors"] == []
    assert payload["error_count"] == 0
    assert payload["test_count"] == 0
    confirm = client.post(f"/api/imports/{payload['token']}/confirm", json={"project_id": current["id"], "supplier_id": "supplier", "batch_name": "batch"})
    assert confirm.status_code == 200
    assert confirm.json()["content_count"] == 1

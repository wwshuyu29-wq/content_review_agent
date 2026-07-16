from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from server.models import AuditRun, Batch, ContentItem, ContentVersion, Issue, ReviewTask, TestCase, TestEvidence
from server.services.excel_import_service import NEW_CONTENT_COLUMNS
from server.services.severity_service import highest_severity

EXPORT_COLUMNS = (
    "系统内容编号",
    "批次编号",
    "格式校验",
    "审核状态",
    "发布状态",
    "问题数量",
    "最高风险等级",
    "问题分类",
    "命中规则",
    "问题原因",
    "原文证据",
    "修改建议",
    "最终标题",
    "最终正文",
    "最终版本来源",
    "是否需要人工",
    "是否已人工确认",
    "审核模型",
    "规则版本",
    "审核完成时间",
    "测试数量", "证据数量", "证据状态", "测试摘要",
    "基础校对维度决策", "基础校对维度分数", "基础校对维度摘要",
    "合规维度决策", "合规维度分数", "合规维度摘要",
    "品牌维度决策", "品牌维度分数", "品牌维度摘要",
    "产品准确维度决策", "产品准确维度分数", "产品准确维度摘要",
    "传播效果维度决策", "传播效果维度分数", "传播效果维度摘要",
)

_MANUAL_SEVERITIES = {"mid", "medium", "high", "unknown", "critical"}
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
_LONG_TEXT_COLUMNS = {
    "正文",
    "内容",
    "问题分类",
    "命中规则",
    "问题原因",
    "原文证据",
    "修改建议",
    "最终标题",
    "最终正文",
}


def export_batch(session: Session, batch_id: int) -> bytes:
    batch = session.scalar(
        select(Batch)
        .where(Batch.id == batch_id)
        .options(
            selectinload(Batch.content_items).selectinload(ContentItem.versions),
            selectinload(Batch.content_items)
            .selectinload(ContentItem.audit_runs)
            .selectinload(AuditRun.issues),
            selectinload(Batch.content_items)
            .selectinload(ContentItem.audit_runs)
            .selectinload(AuditRun.rule_version),
            selectinload(Batch.content_items).selectinload(ContentItem.test_cases).selectinload(TestCase.evidence).selectinload(TestEvidence.asset),
            selectinload(Batch.content_items)
            .selectinload(ContentItem.review_tasks)
            .selectinload(ReviewTask.human_decisions),
        )
    )
    if batch is None:
        raise ValueError(f"Batch {batch_id} does not exist")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "批次导出"
    headers = list(NEW_CONTENT_COLUMNS) + list(EXPORT_COLUMNS)
    worksheet.append([_sanitize_excel_value(value) for value in headers])

    for item in sorted(batch.content_items, key=lambda content: content.id):
        worksheet.append([_sanitize_excel_value(value) for value in _row_for_item(batch, item)])

    _style_worksheet(worksheet, headers)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _sanitize_excel_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _row_for_item(batch: Batch, item: ContentItem) -> list[Any]:
    supplier_version = _supplier_version(item)
    latest_version = _latest_version(item)
    payload = dict(supplier_version.payload) if supplier_version is not None else {}
    latest_audit = _latest_audit(item)
    issues = _sorted_issues(latest_audit.issues) if latest_audit is not None else []
    tasks = _tasks_for_audit(item, latest_audit)

    return [
        _payload_value(payload, "title", default=supplier_version.title if supplier_version is not None else item.title),
        _payload_value(payload, "body", default=supplier_version.body if supplier_version is not None else ""),
        _payload_value(payload, "account_type"),
        _payload_value(payload, "platform"),
        _payload_value(payload, "account_name"),
        _payload_value(payload, "publish_time"),
        _payload_value(payload, "image_filename"),
        item.id,
        batch.id,
        _enum_value(item.format_status),
        _enum_value(item.review_status),
        _enum_value(item.publish_status),
        len(issues),
        _highest_severity(issues),
        _join_issue_field(issues, "category"),
        _join_issue_field(issues, "rule_id"),
        _join_issue_field(issues, "reason"),
        _join_issue_field(issues, "evidence_quote"),
        _join_issue_field(issues, "suggestion"),
        latest_version.title if latest_version is not None else "",
        latest_version.body if latest_version is not None else "",
        latest_version.source if latest_version is not None else "",
        _yes_no(_needs_human(issues, tasks)),
        _yes_no(_human_confirmed(tasks)),
        latest_audit.model if latest_audit is not None else "",
        latest_audit.rule_version.version if latest_audit is not None and latest_audit.rule_version is not None else "",
        _excel_datetime(latest_audit.completed_at if latest_audit is not None else None),
        len(item.test_cases), len({binding.asset_id for test in item.test_cases for binding in test.evidence}),
        "PRESENT" if item.test_cases and all(test.evidence for test in item.test_cases) else ("MISSING" if item.test_cases else "NONE"),
        "\n".join(f"{test.external_test_case_id}: {test.observed_result}" for test in item.test_cases),
        *sum(([result.decision or "", result.score, result.summary or ""] for result in _agent_results(latest_audit)), []),
    ]


def _agent_results(audit: Optional[AuditRun]) -> list[Any]:
    wanted = ["CONTENT_QUALITY", "COMPLIANCE", "BRAND", "PRODUCT_ACCURACY", "CAMPAIGN_EFFECTIVENESS"]
    values = {result.agent_id or result.agent_name: result for result in (audit.agent_results if audit else [])}
    return [values.get(name, _EmptyAgent()) for name in wanted]


class _EmptyAgent:
    decision = ""
    score = None
    summary = ""


def _supplier_version(item: ContentItem) -> Optional[ContentVersion]:
    if not item.versions:
        return None
    return min(item.versions, key=lambda version: version.version)


def _latest_version(item: ContentItem) -> Optional[ContentVersion]:
    if not item.versions:
        return None
    return max(item.versions, key=lambda version: version.version)


def _latest_audit(item: ContentItem) -> Optional[AuditRun]:
    if not item.audit_runs:
        return None
    return max(item.audit_runs, key=lambda audit: audit.id or 0)


def _tasks_for_audit(item: ContentItem, audit: Optional[AuditRun]) -> list[ReviewTask]:
    if audit is None:
        return []
    return sorted(
        [task for task in item.review_tasks if task.audit_run_id == audit.id],
        key=lambda task: task.id or 0,
    )


def _sorted_issues(issues: Iterable[Issue]) -> list[Issue]:
    return sorted(issues, key=lambda issue: issue.id or 0)


def _payload_value(payload: dict[str, Any], key: str, default: Any = "") -> Any:
    value = payload.get(key, default)
    return default if value is None else value


def _enum_value(value: Any) -> str:
    return value.value if hasattr(value, "value") else str(value)


def _join_issue_field(issues: list[Issue], field: str) -> str:
    return "\n".join(str(getattr(issue, field) or "") for issue in issues)


def _highest_severity(issues: list[Issue]) -> str:
    return highest_severity((issue.severity for issue in issues), empty="") or ""


def _needs_human(issues: list[Issue], tasks: list[ReviewTask]) -> bool:
    if tasks:
        return True
    return any(issue.human_required or issue.severity.lower() in _MANUAL_SEVERITIES for issue in issues)


def _human_confirmed(tasks: list[ReviewTask]) -> bool:
    if not tasks or any(task.status == "OPEN" for task in tasks):
        return False
    return any(task.human_decisions for task in tasks)


def _yes_no(value: bool) -> str:
    return "是" if value else "否"


def _excel_datetime(value: Optional[datetime]) -> Optional[datetime]:
    if value is not None and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _style_worksheet(worksheet, headers: list[str]) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = wrap_top

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_top

    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        if header in {"内容", "正文", "最终正文"}:
            width = 48
        elif header in _LONG_TEXT_COLUMNS:
            width = 32
        else:
            width = max(12, min(24, len(header) + 6))
        worksheet.column_dimensions[letter].width = width

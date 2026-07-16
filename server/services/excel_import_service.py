from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import stat
import threading
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path, PurePosixPath
from secrets import token_urlsafe
from typing import Any, Dict, List, Optional, Tuple
from zipfile import BadZipFile, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from server.models import Batch, FormatStatus, Project
from server.services.content_service import MAX_BODY_LENGTH, MAX_TITLE_LENGTH, submit_batch


IMPORT_COLUMNS = (
    "供应商内容编号",
    "活动主题",
    "平台",
    "标题",
    "正文",
    "计划发布时间",
    "备注",
)
CONTENT_COLUMNS = ("供应商内容编号", "活动主题", "账号名称", "账号类型", "平台", "标题", "正文", "计划发布时间", "备注")
NEW_CONTENT_COLUMNS = ("标题", "内容", "类型", "目标平台", "作者", "发布日期")
OPTIMIZED_CONTENT_COLUMNS = ("标题", "内容", "类型", "目标平台", "作者", "发布日期", "优化后版本")
TEST_CASE_COLUMNS = ("供应商内容编号", "测试场景编号", "测试结论", "测试指令", "实际返回结果", "测试城市", "测试时间", "百度地图版本", "设备", "操作系统", "网络环境", "证据文件名")
REQUIRED_CONTENT_COLUMNS = CONTENT_COLUMNS[:7]
EVIDENCE_SUFFIXES = frozenset({".png", ".jpg", ".jpeg", ".webp", ".gif", ".mp4", ".mov", ".webm", ".txt", ".log", ".json"})
MAX_EVIDENCE_BYTES = 100 * 1024 * 1024
REQUIRED_COLUMNS = IMPORT_COLUMNS[:5]
MAX_IMPORT_ROWS = 500
EXCEL_CELL_TEXT_LIMIT = 32_767
MAX_XLSX_ENTRIES = 2000
MAX_XLSX_ENTRY_BYTES = 50 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 200 * 1024 * 1024
MAX_REVIEW_BRIEF_LENGTH = 10_000
DEFAULT_EVIDENCE_TRIGGER_TERMS = ("亲测", "实测", "自用")
PREVIEW_TTL = timedelta(hours=2)
PREVIEW_MANIFEST_VERSION = 3
PREVIEW_ROOT_REGISTRY_ENV = "CONTENT_REVIEW_PREVIEW_ROOT_REGISTRY"
DEFAULT_PREVIEW_ROOT_REGISTRY = Path(__file__).resolve().parents[2] / "data" / "preview-roots.json"
_COLUMN_KEYS = {
    "供应商内容编号": "external_id",
    "活动主题": "campaign_theme",
    "平台": "platform",
    "标题": "title",
    "正文": "body",
    "计划发布时间": "publish_time",
    "备注": "note",
    "账号名称": "account_name",
    "账号类型": "account_type",
    "内容": "body",
    "类型": "account_type",
    "目标平台": "platform",
    "作者": "account_name",
    "发布日期": "publish_time",
    "优化后版本": "body",
}
_TEST_COLUMN_KEYS = {"供应商内容编号": "供应商内容编号", "测试场景编号": "测试场景编号", "测试结论": "测试结论", "测试指令": "测试指令", "实际返回结果": "实际返回结果", "测试城市": "测试城市", "测试时间": "测试时间", "百度地图版本": "百度地图版本", "设备": "设备", "操作系统": "操作系统", "网络环境": "网络环境", "证据文件名": "证据文件名"}
_WINDOWS_DRIVE = re.compile(r"^[A-Za-z]:")
_TOKEN_PATTERN = re.compile(r"^[A-Za-z0-9_-]{32,128}$")
_PREVIEW_KEYS = frozenset(
    {
        "version",
        "token",
        "expires_at",
        "rows",
        "warnings",
        "total_count",
        "valid_count",
        "error_count",
        "test_cases",
        "errors",
        "project_id", "project_code", "content_type", "package_version",
        "supplier_id", "batch_name", "project_type", "owner_name",
        "review_brief", "brief_summary",
    }
)
_ROW_KEYS = frozenset({"manuscript_index", "row_number", "normalized", "errors", "warnings", "valid", "tests"})
_NORMALIZED_KEYS = frozenset(_COLUMN_KEYS.values())


@dataclass(frozen=True)
class PreviewIdentity:
    project_id: int
    project_code: str
    content_type: str
    package_version: str
    supplier_id: str
    batch_name: str
    project_type: str
    owner_name: str


@dataclass(frozen=True)
class TestCasePreview:
    content_external_id: str
    external_test_case_id: str
    claim: str
    command: str
    observed_result: str
    city: Optional[str]
    tested_at: Optional[str]
    app_version: Optional[str]
    device: Optional[str]
    operating_system: Optional[str]
    network_environment: Optional[str]
    evidence_filenames: List[str]

@dataclass(frozen=True)
class ImportRowPreview:
    manuscript_index: int
    row_number: int
    normalized: Dict[str, Any]
    errors: List[str]
    warnings: List[str]
    valid: bool
    tests: List[TestCasePreview] = field(default_factory=list)


@dataclass(frozen=True)
class ImportPreview:
    token: str
    rows: List[ImportRowPreview]
    warnings: List[str]
    total_count: int
    valid_count: int
    error_count: int
    review_brief: str
    brief_summary: str
    test_cases: List[TestCasePreview] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    identity: Optional[PreviewIdentity] = None

    @property
    def test_count(self) -> int:
        return len(self.test_cases or [])


@dataclass(frozen=True)
class _PreviewLocation:
    temp_root: Path
    preview_dir: Path
    manifest: Path


_preview_locations: Dict[str, _PreviewLocation] = {}
_preview_locations_lock = threading.Lock()


def build_import_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "内容清单"
    worksheet.append(list(NEW_CONTENT_COLUMNS))
    workbook.create_sheet("字段说明").append(["本表仅用于说明字段，导入时忽略"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def preview_import(
    xlsx_path: Path,
    zip_path: Optional[Path],
    temp_root: Path,
    *,
    review_brief: str = "测试批次 Brief",
    trigger_terms: Tuple[str, ...] = DEFAULT_EVIDENCE_TRIGGER_TERMS,
    identity: Optional[PreviewIdentity] = None,
) -> ImportPreview:
    token = token_urlsafe(32)
    root = temp_root.resolve()
    preview_dir = root / token
    manifest = preview_dir / "preview.json"
    registered = False

    try:
        if zip_path is not None:
            raise ValueError("当前仅支持文字审核，不支持媒体 ZIP 或证据 ZIP")
        preview_dir.mkdir(parents=True, exist_ok=False)
        brief = _normalize_review_brief(review_brief)
        rows, test_cases = _read_workbook(xlsx_path)
        warnings: List[str] = []
        rows, test_cases, preview_errors = _validate_test_cases(rows, test_cases, trigger_terms)
        rows = _mark_duplicate_external_ids(rows)

        valid_count = sum(row.valid for row in rows)
        preview = ImportPreview(
            token=token,
            rows=rows,
            warnings=warnings,
            total_count=len(rows),
            valid_count=valid_count,
            error_count=len(rows) - valid_count + len(preview_errors),
            review_brief=brief,
            brief_summary=_brief_summary(brief),
            test_cases=test_cases,
            errors=preview_errors,
            identity=identity,
        )
        expires_at = datetime.now(timezone.utc) + PREVIEW_TTL
        _write_preview(manifest, preview, expires_at)
        location = _PreviewLocation(root, preview_dir.resolve(), manifest.resolve())
        with _preview_locations_lock:
            _register_preview_root(root)
            _preview_locations[token] = location
            registered = True
        cleanup_expired_previews()
        return preview
    except Exception:
        if registered:
            with _preview_locations_lock:
                _preview_locations.pop(token, None)
        shutil.rmtree(preview_dir, ignore_errors=True)
        raise


def load_preview(token: str) -> ImportPreview:
    location = _resolve_preview_location(token)
    try:
        preview, expires_at = _load_preview_manifest(location, token)
    except _ExpiredPreviewError as exc:
        _remove_preview(token, location)
        raise ValueError("导入预览已过期") from exc
    return preview


def consume_preview(token: str) -> ImportPreview:
    location = _resolve_preview_location(token)
    consumed_dir = location.preview_dir.with_name(location.preview_dir.name + ".consuming")
    with _preview_locations_lock:
        try:
            location.preview_dir.rename(consumed_dir)
        except FileNotFoundError as exc:
            _preview_locations.pop(token, None)
            raise ValueError("导入 token 不存在或已失效") from exc
        except OSError as exc:
            raise ValueError("导入预览无法消费") from exc
        _preview_locations.pop(token, None)

    consumed_location = _PreviewLocation(location.temp_root, consumed_dir, consumed_dir / "preview.json")
    try:
        preview, _ = _load_preview_manifest(consumed_location, token)
        return preview
    finally:
        shutil.rmtree(consumed_dir, ignore_errors=True)


def confirm_import(
    session: Session,
    token: str,
    project_id: int,
    supplier_id: str,
    batch_name: str,
    *,
    project_type: str | None = None,
    owner_name: str | None = None,
    uploaded_by_user_id: int | None = None,
) -> Batch:
    if not isinstance(token, str) or not _TOKEN_PATTERN.fullmatch(token):
        raise ValueError("无效的导入 token")

    existing = session.scalar(select(Batch).where(Batch.import_token == token))
    if existing is not None:
        identity_bound = any(
            version.payload.get("preview_identity")
            for item in existing.content_items for version in item.versions[:1]
        )
        normalized_project_type = project_type.strip() if isinstance(project_type, str) and project_type.strip() else existing.project_type
        normalized_owner_name = owner_name.strip() if isinstance(owner_name, str) and owner_name.strip() else existing.owner_name
        if identity_bound and (
            existing.project_id, existing.supplier_id, existing.name, existing.project_type, existing.owner_name,
        ) != (project_id, supplier_id, batch_name, normalized_project_type, normalized_owner_name):
            raise ValueError("导入确认信息与原批次身份不匹配")
        return existing

    location = _resolve_preview_location(token)
    try:
        preview, _ = _load_preview_manifest(location, token)
    except _ExpiredPreviewError as exc:
        _remove_preview(token, location)
        raise ValueError("导入预览已过期") from exc

    if preview.errors:
        raise ValueError("导入预览包含全局错误，无法确认")
    if preview.identity is not None:
        project = session.get(Project, project_id)
        if project is None:
            raise ValueError(f"Project {project_id} does not exist")
        expected = preview.identity
        current_package = project.current_rule_version.package_version if project.current_rule_version else None
        normalized_project_type = (project_type or project.name or "").strip()
        normalized_owner_name = (owner_name or supplier_id).strip()
        actual = (
            project_id, project.code, project.content_type, current_package,
            supplier_id, batch_name, normalized_project_type, normalized_owner_name,
        )
        bound = (
            expected.project_id, expected.project_code, expected.content_type, expected.package_version,
            expected.supplier_id, expected.batch_name, expected.project_type, expected.owner_name,
        )
        if actual != bound:
            raise ValueError("导入确认信息与预览身份不匹配")
    confirmed_project_type = preview.identity.project_type if preview.identity is not None else project_type
    confirmed_owner_name = preview.identity.owner_name if preview.identity is not None else owner_name

    commit_completed = False
    try:
        contents = _build_confirm_contents(preview)
        batch = submit_batch(
            session,
            project_id=project_id,
            supplier_id=supplier_id,
            name=batch_name,
            contents=contents,
            import_token=token,
            review_brief=preview.review_brief,
            project_type=confirmed_project_type,
            owner_name=confirmed_owner_name,
            uploaded_by_user_id=uploaded_by_user_id,
            commit=False,
        )
        session.commit()
        commit_completed = True
    except IntegrityError:
        if not commit_completed:
            session.rollback()
            existing = session.scalar(select(Batch).where(Batch.import_token == token))
            if existing is not None:
                if preview.identity is not None and (
                    existing.project_id, existing.supplier_id, existing.name, existing.project_type, existing.owner_name,
                ) != (project_id, supplier_id, batch_name, preview.identity.project_type, preview.identity.owner_name):
                    raise ValueError("并发导入批次身份与预览不匹配")
                return existing
        raise
    except Exception:
        if not commit_completed:
            session.rollback()
        raise

    session.refresh(batch)
    _remove_preview(token, location)
    return batch


class _ExpiredPreviewError(ValueError):
    pass


def _build_confirm_contents(preview: ImportPreview) -> List[Dict[str, Any]]:
    contents: List[Dict[str, Any]] = []
    for row in preview.rows:
        normalized = row.normalized
        payload = _payload_for_row(preview.token, row)
        if preview.identity is not None:
            payload["preview_identity"] = asdict(preview.identity)
        contents.append(
            {
                "external_id": _external_id_for_row(preview.token, row),
                "title": normalized.get("title") or "",
                "body": normalized.get("body") or "",
                "payload": payload,
                "format_status": _format_status_for_row(row),
            }
        )
    return contents


def _payload_for_row(token: str, row: ImportRowPreview) -> Dict[str, Any]:
    normalized = row.normalized
    return {
        "supplier_external_id": normalized.get("external_id"),
        "campaign_theme": normalized.get("campaign_theme"),
        "account_name": normalized.get("account_name"),
        "account_type": normalized.get("account_type"),
        "platform": normalized.get("platform"),
        "title": normalized.get("title"),
        "body": normalized.get("body"),
        "publish_time": normalized.get("publish_time"),
        "note": normalized.get("note"),
        "manuscript_index": row.manuscript_index,
        "row_number": row.row_number,
        "preview_errors": list(row.errors),
        "preview_warnings": list(row.warnings),
        "import_token": token,
    }


def _external_id_for_row(token: str, row: ImportRowPreview) -> str:
    external_id = row.normalized.get("external_id")
    if row.valid and external_id:
        return external_id
    return f"import:{token[:16]}:row:{row.row_number}"


def _format_status_for_row(row: ImportRowPreview) -> FormatStatus:
    if row.valid:
        return FormatStatus.PASSED
    required_values = (
        row.normalized.get("external_id"),
        row.normalized.get("campaign_theme"),
        row.normalized.get("platform"),
        row.normalized.get("title"),
        row.normalized.get("body"),
    )
    if any(not value for value in required_values):
        return FormatStatus.INCOMPLETE
    return FormatStatus.INVALID


def _resolve_preview_location(token: str) -> _PreviewLocation:
    if not isinstance(token, str) or not _TOKEN_PATTERN.fullmatch(token):
        raise ValueError("无效的导入 token")

    with _preview_locations_lock:
        location = _preview_locations.get(token)
        if location is None:
            location = _find_persisted_preview(token)
            if location is not None:
                _preview_locations[token] = location
    if location is None:
        raise ValueError("导入 token 不存在或已失效")
    return location


def _find_persisted_preview(token: str) -> Optional[_PreviewLocation]:
    for root in _load_preview_roots():
        preview_dir = root / token
        manifest = preview_dir / "preview.json"
        if preview_dir.is_dir() and manifest.is_file():
            return _PreviewLocation(root, preview_dir, manifest)
    return None


def _load_preview_manifest(
    location: _PreviewLocation, token: str
) -> Tuple[ImportPreview, datetime]:
    try:
        preview_dir = location.preview_dir.resolve(strict=True)
        manifest = location.manifest.resolve(strict=True)
        root = location.temp_root.resolve(strict=True)
    except (FileNotFoundError, RuntimeError) as exc:
        raise ValueError("导入预览不存在或已失效") from exc

    if preview_dir.parent != root or manifest.parent != preview_dir or not manifest.is_file():
        raise ValueError("导入预览路径无效")

    try:
        payload = json.loads(manifest.read_text(encoding="utf-8"))
        preview, expires_at = _preview_from_dict(payload)
        _validate_preview_files(location, preview)
    except (OSError, TypeError, KeyError, ValueError, json.JSONDecodeError) as exc:
        raise ValueError("导入预览数据无效") from exc
    if preview.token != token:
        raise ValueError("导入预览 token 不匹配")
    if expires_at <= datetime.now(timezone.utc):
        raise _ExpiredPreviewError("导入预览已过期")
    return preview, expires_at


def _validate_preview_files(location: _PreviewLocation, preview: ImportPreview) -> None:
    evidence_root = location.preview_dir / "evidence"
    valid_owners = {row.normalized.get("external_id") for row in preview.rows if row.valid}
    for test in preview.test_cases:
        if test.content_external_id not in valid_owners:
            continue
        for filename in test.evidence_filenames:
            path = evidence_root / filename
            try:
                root = evidence_root.resolve(strict=True)
                resolved = path.resolve(strict=True)
            except (FileNotFoundError, RuntimeError) as exc:
                raise ValueError(f"预览证据文件不存在：{filename}") from exc
            if resolved.parent != root or not resolved.is_file() or resolved.stat().st_size > MAX_EVIDENCE_BYTES:
                raise ValueError(f"预览证据文件无效：{filename}")

def _remove_preview(token: str, location: _PreviewLocation) -> None:
    with _preview_locations_lock:
        _preview_locations.pop(token, None)
    shutil.rmtree(location.preview_dir, ignore_errors=True)


def _preview_root_registry_path() -> Path:
    configured = os.environ.get(PREVIEW_ROOT_REGISTRY_ENV)
    return Path(configured).expanduser().resolve() if configured else DEFAULT_PREVIEW_ROOT_REGISTRY.resolve()


def _register_preview_root(root: Path) -> None:
    registry = _preview_root_registry_path()
    roots = _load_preview_roots()
    if root in roots:
        return
    roots.append(root)
    registry.parent.mkdir(parents=True, exist_ok=True)
    _write_json_atomic(registry, {"version": 1, "roots": sorted(str(path) for path in roots)})


def _load_preview_roots() -> List[Path]:
    registry = _preview_root_registry_path()
    if not registry.exists():
        return []
    try:
        payload = json.loads(registry.read_text(encoding="utf-8"))
        if not isinstance(payload, dict) or set(payload) != {"version", "roots"}:
            raise ValueError
        if type(payload["version"]) is not int or payload["version"] != 1:
            raise ValueError
        if not isinstance(payload["roots"], list) or not all(
            isinstance(root, str) and Path(root).is_absolute() for root in payload["roots"]
        ):
            raise ValueError
        roots = [Path(root).resolve() for root in payload["roots"]]
        if len(roots) != len(set(roots)):
            raise ValueError
        return roots
    except (OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise ValueError("导入预览根目录配置无效") from exc


def _validate_xlsx_archive(xlsx_path: Path) -> None:
    required = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
    total = 0
    names = set()
    try:
        with ZipFile(xlsx_path) as archive:
            if len(archive.filelist) > MAX_XLSX_ENTRIES:
                raise ValueError("XLSX 条目数量超过安全限制")
            for info in archive.filelist:
                _validate_zip_path(info.filename)
                if info.flag_bits & 0x1:
                    raise ValueError(f"XLSX 包含加密条目：{info.filename}")
                if _is_zip_symlink(info):
                    raise ValueError(f"XLSX 不允许符号链接：{info.filename}")
                if info.filename in names:
                    raise ValueError(f"XLSX 条目重复：{info.filename}")
                if info.file_size > MAX_XLSX_ENTRY_BYTES:
                    raise ValueError(f"XLSX 单条目超过安全限制：{info.filename}")
                total += info.file_size
                if total > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("XLSX 解压后内容超过安全限制")
                names.add(info.filename)
    except (BadZipFile, OSError) as exc:
        raise ValueError("Excel 文件无法读取") from exc
    missing = sorted(required - names)
    if missing:
        raise ValueError("XLSX 缺少必需 OOXML 部件：" + "、".join(missing))


def _read_workbook(xlsx_path: Path) -> Tuple[List[ImportRowPreview], List[TestCasePreview]]:
    workbook = None
    try:
        _validate_xlsx_archive(xlsx_path)
        workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
        if "内容清单" not in workbook.sheetnames:
            raise ValueError("Excel 缺少命名工作表：内容清单")
        rows = _read_sheet_rows(workbook["内容清单"])
        tests: List[TestCasePreview] = []
        return rows, tests
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("Excel 工作表解析失败") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _read_sheet_rows(sheet) -> List[ImportRowPreview]:
    iterator = sheet.iter_rows()
    header_cells = next(iterator, None)
    if header_cells is None:
        raise ValueError("Excel 表头不能为空")
    if any(cell.data_type == "f" for cell in header_cells):
        raise ValueError("Excel 表头不允许公式")
    headers = _validate_headers(tuple(cell.value for cell in header_cells), ())
    if headers == NEW_CONTENT_COLUMNS:
        format_name = "new"
    elif headers == OPTIMIZED_CONTENT_COLUMNS:
        format_name = "new_optimized"
    elif any(column in headers for column in NEW_CONTENT_COLUMNS[1:] + ("优化后版本",)):
        raise ValueError("Excel 新模板表头必须严格匹配：" + "、".join(NEW_CONTENT_COLUMNS) + "；也支持将最后一列替换为：优化后版本")
    else:
        required = REQUIRED_CONTENT_COLUMNS if all(column in headers for column in ("账号名称", "账号类型")) else REQUIRED_COLUMNS
        _validate_headers(headers, required)
        format_name = "named_legacy" if required == REQUIRED_CONTENT_COLUMNS else "legacy"
    indexes = {header: index for index, header in enumerate(headers)}
    rows = []
    for row_number, cells in enumerate(iterator, 2):
        values = tuple(cell.value for cell in cells)
        if _is_blank_row(values):
            continue
        if len(rows) >= MAX_IMPORT_ROWS:
            raise ValueError("Excel 最多允许 500 条内容")
        formulas = {index for index, cell in enumerate(cells) if cell.data_type == "f"}
        rows.append(_normalize_row(len(rows) + 1, row_number, values, indexes, headers, formulas, format_name=format_name))
    return rows


def _split_filenames(value: Optional[str]) -> List[str]:
    return [part.strip() for part in re.split(r"[,，;；]", value or "") if part.strip()]


def _group_tests(tests: List[TestCasePreview]) -> Dict[str, List[TestCasePreview]]:
    grouped: Dict[str, List[TestCasePreview]] = {}
    for test in tests:
        grouped.setdefault(test.content_external_id or "", []).append(test)
    return grouped


def _validate_test_cases(rows, tests, _trigger_terms):
    content_ids = {
        row.normalized.get("external_id")
        for row in rows
        if row.valid and row.normalized.get("external_id")
    }
    accepted_tests: List[TestCasePreview] = []
    seen = set()
    for test in tests:
        test_id = test.external_test_case_id
        if (
            test.content_external_id not in content_ids
            or not test_id
            or test_id in seen
            or not test.command
            or not test.observed_result
        ):
            continue
        seen.add(test_id)
        accepted_tests.append(test)

    grouped = _group_tests(accepted_tests)
    updated = [
        ImportRowPreview(
            row.manuscript_index,
            row.row_number,
            row.normalized,
            list(row.errors),
            row.warnings,
            row.valid,
            grouped.get(row.normalized.get("external_id") or "", []),
        )
        for row in rows
    ]
    return updated, accepted_tests, []


def _validate_headers(raw_headers: Tuple[Any, ...], required_columns=None) -> Tuple[str, ...]:
    headers: List[str] = []
    for value in raw_headers:
        if value is None or not str(value).strip():
            raise ValueError("Excel 表头不得为空")
        headers.append(str(value).strip())

    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise ValueError("Excel 表头重复：" + "、".join(duplicates))

    required = REQUIRED_COLUMNS if required_columns is None else required_columns
    missing = [column for column in required if column not in headers]
    if missing:
        raise ValueError("Excel 缺少必需表头：" + "、".join(missing))
    return tuple(headers)


def _is_blank_row(values: Tuple[Any, ...]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _normalize_row(
    manuscript_index: int,
    row_number: int,
    values: Tuple[Any, ...],
    indexes: Dict[str, int],
    headers: Tuple[str, ...],
    formula_indexes: set[int],
    format_name: str = "legacy",
) -> ImportRowPreview:
    normalized: Dict[str, Any] = {}
    errors = [
        f"第 {row_number} 行 {headers[index]} 不允许公式"
        for index in sorted(formula_indexes)
        if index < len(headers)
    ]

    columns = (
        NEW_CONTENT_COLUMNS if format_name == "new"
        else OPTIMIZED_CONTENT_COLUMNS if format_name == "new_optimized"
        else CONTENT_COLUMNS if format_name == "named_legacy"
        else IMPORT_COLUMNS
    )
    normalized = {key: None for key in _NORMALIZED_KEYS} if format_name in {"new", "new_optimized"} else {}
    for column in columns:
        index = indexes.get(column)
        raw_value = (
            values[index]
            if index is not None and index < len(values) and index not in formula_indexes
            else None
        )
        key = _COLUMN_KEYS[column]
        if column in {"计划发布时间", "发布日期"}:
            normalized[key], date_error = _normalize_date(raw_value, column)
            if date_error:
                errors.append(date_error)
        else:
            normalized[key] = _normalize_text(raw_value)

    if format_name == "new_optimized":
        content_index = indexes.get("内容")
        optimized_index = indexes.get("优化后版本")
        original_body = _normalize_text(values[content_index]) if content_index is not None and content_index < len(values) and content_index not in formula_indexes else None
        optimized_body = _normalize_text(values[optimized_index]) if optimized_index is not None and optimized_index < len(values) and optimized_index not in formula_indexes else None
        normalized["body"] = optimized_body or original_body
        if optimized_body and original_body and optimized_body != original_body:
            normalized["note"] = "导入时使用“优化后版本”作为审核正文"

    required_columns = ("标题", "内容", "目标平台") if format_name in {"new", "new_optimized"} else (REQUIRED_CONTENT_COLUMNS if format_name == "named_legacy" else REQUIRED_COLUMNS)
    for column in required_columns:
        if not normalized[_COLUMN_KEYS[column]]:
            errors.append(f"{column}不能为空")

    if format_name in {"new", "new_optimized"}:
        normalized["external_id"] = _derive_external_id(row_number, normalized["title"], normalized["account_name"])

    external_id = normalized["external_id"]
    if external_id and len(external_id) > 200:
        errors.append("供应商内容编号不能超过 200 个字符")
    title = normalized["title"]
    if title and len(title) > MAX_TITLE_LENGTH:
        errors.append(f"标题不能超过 {MAX_TITLE_LENGTH} 个字符")
    body = normalized["body"]
    if body and len(body) >= EXCEL_CELL_TEXT_LIMIT:
        errors.append("正文达到 Excel 单元格上限，可能已被截断")
    elif body and len(body) > MAX_BODY_LENGTH:
        errors.append(f"正文不能超过 {MAX_BODY_LENGTH} 个字符")

    return ImportRowPreview(manuscript_index, row_number, normalized, errors, [], not errors, [])


def _normalize_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _derive_external_id(row_number: int, title: Optional[str], author: Optional[str]) -> str:
    source = json.dumps([title or "", author or ""], ensure_ascii=False, separators=(",", ":"))
    digest = hashlib.sha256(source.encode("utf-8")).hexdigest()[:20]
    return f"excel:{row_number}:{digest}"


def _normalize_date(value: Any, column: str = "计划发布时间") -> Tuple[Optional[str], Optional[str]]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, None
    if isinstance(value, datetime):
        return value.date().isoformat(), None
    if isinstance(value, date):
        return value.isoformat(), None
    text = str(value).strip()
    if re.fullmatch(r"\d+(\.\d+)?", text):
        try:
            converted = from_excel(float(text))
            if isinstance(converted, datetime):
                return converted.date().isoformat(), None
            if isinstance(converted, date):
                return converted.isoformat(), None
        except (TypeError, ValueError, OverflowError):
            pass
    try:
        return date.fromisoformat(text).isoformat(), None
    except ValueError:
        return text, f"{column}必须为 YYYY-MM-DD 或 Excel 日期"


def _validate_zip_path(name: str) -> None:
    if not name or "\\" in name or name.startswith("/") or _WINDOWS_DRIVE.match(name):
        raise ValueError(f"ZIP 包含不安全路径：{name}")
    path = PurePosixPath(name)
    if path.is_absolute() or any(part in {"", ".", ".."} for part in path.parts):
        raise ValueError(f"ZIP 包含不安全路径：{name}")


def _is_zip_symlink(info: ZipInfo) -> bool:
    return info.create_system == 3 and stat.S_ISLNK(info.external_attr >> 16)


def _is_safe_basename(name: str) -> bool:
    if not name or name in {".", ".."} or "/" in name or "\\" in name:
        return False
    if name.startswith("/") or _WINDOWS_DRIVE.match(name):
        return False
    return PurePosixPath(name).name == name


def _mark_duplicate_external_ids(rows: List[ImportRowPreview]) -> List[ImportRowPreview]:
    counts: Dict[str, int] = {}
    for row in rows:
        external_id = row.normalized["external_id"]
        if external_id:
            counts[external_id] = counts.get(external_id, 0) + 1

    updated = []
    for row in rows:
        errors = list(row.errors)
        external_id = row.normalized["external_id"]
        if external_id and counts[external_id] > 1:
            errors.append(f"内容编号在批次内重复：{external_id}")
        updated.append(_replace_row_errors(row, errors))
    return updated


def _replace_row_errors(row: ImportRowPreview, errors: List[str]) -> ImportRowPreview:
    return ImportRowPreview(row.manuscript_index, row.row_number, row.normalized, errors, row.warnings, not errors, row.tests)


def _write_preview(path: Path, preview: ImportPreview, expires_at: datetime) -> None:
    payload = _preview_to_dict(preview)
    payload["version"] = PREVIEW_MANIFEST_VERSION
    payload["expires_at"] = expires_at.isoformat()
    payload["test_cases"] = [_test_to_dict(test) for test in (preview.test_cases or [])]
    identity = preview.identity
    for key in ("project_id", "project_code", "content_type", "package_version", "supplier_id", "batch_name", "project_type", "owner_name"):
        payload[key] = getattr(identity, key) if identity is not None else None
    _write_json_atomic(path, payload)


def _write_json_atomic(path: Path, payload: Dict[str, Any]) -> None:
    temporary = path.with_name(f".{path.name}.{token_urlsafe(8)}.tmp")
    try:
        with temporary.open("x", encoding="utf-8") as output:
            json.dump(payload, output, ensure_ascii=False, separators=(",", ":"))
            output.flush()
            os.fsync(output.fileno())
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)


def _normalize_review_brief(value: str) -> str:
    if not isinstance(value, str):
        raise ValueError("Brief 内容无效")
    normalized = "\n".join(line.rstrip() for line in value.replace("\r\n", "\n").replace("\r", "\n").split("\n")).strip()
    if not normalized:
        raise ValueError("请上传或填写本批次 Brief")
    if len(normalized) > MAX_REVIEW_BRIEF_LENGTH:
        raise ValueError("Brief 不能超过 10000 个字符")
    return normalized


def _brief_summary(value: str) -> str:
    compact = re.sub(r"\s+", " ", value).strip()
    return compact[:160] + ("…" if len(compact) > 160 else "")


def _preview_to_dict(preview: ImportPreview) -> Dict[str, Any]:
    payload = asdict(preview)
    payload.pop("identity", None)
    return payload


def _test_to_dict(test: TestCasePreview) -> Dict[str, Any]:
    return {"content_external_id": test.content_external_id, "external_test_case_id": test.external_test_case_id, "claim": test.claim, "command": test.command, "observed_result": test.observed_result, "city": test.city, "tested_at": test.tested_at, "app_version": test.app_version, "device": test.device, "operating_system": test.operating_system, "network_environment": test.network_environment, "evidence_filenames": list(test.evidence_filenames)}

def _test_from_dict(value: Any) -> TestCasePreview:
    if not isinstance(value, dict) or set(value) != {"content_external_id", "external_test_case_id", "claim", "command", "observed_result", "city", "tested_at", "app_version", "device", "operating_system", "network_environment", "evidence_filenames"} or not _is_string_list(value["evidence_filenames"]):
        raise ValueError("测试场景字段无效")
    return TestCasePreview(**value)

def _preview_identity_from_dict(payload: Dict[str, Any]) -> Optional[PreviewIdentity]:
    keys = ("project_id", "project_code", "content_type", "package_version", "supplier_id", "batch_name", "project_type", "owner_name")
    values = [payload[key] for key in keys]
    if all(value is None for value in values):
        return None
    if type(payload["project_id"]) is not int or payload["project_id"] <= 0:
        raise ValueError("预览项目身份无效")
    if not all(isinstance(payload[key], str) and payload[key].strip() for key in keys[1:]):
        raise ValueError("预览项目身份无效")
    return PreviewIdentity(**{key: payload[key] for key in keys})


def _preview_from_dict(payload: Dict[str, Any]) -> Tuple[ImportPreview, datetime]:
    if not isinstance(payload, dict) or set(payload) != _PREVIEW_KEYS:
        raise ValueError("预览字段无效")
    if type(payload["version"]) is not int or payload["version"] != PREVIEW_MANIFEST_VERSION:
        raise ValueError("预览版本无效")
    if not isinstance(payload["token"], str) or not _TOKEN_PATTERN.fullmatch(payload["token"]):
        raise ValueError("预览 token 无效")
    expires_at = _parse_manifest_expiry(payload["expires_at"])
    if not isinstance(payload["rows"], list) or len(payload["rows"]) > MAX_IMPORT_ROWS:
        raise ValueError("预览行数无效")
    if not isinstance(payload["test_cases"], list) or len(payload["test_cases"]) > MAX_IMPORT_ROWS:
        raise ValueError("预览测试场景数量无效")
    if not _is_string_list(payload["warnings"]) or not _is_string_list(payload["errors"]):
        raise ValueError("预览消息无效")
    if not isinstance(payload["review_brief"], str) or not payload["review_brief"].strip():
        raise ValueError("预览 Brief 无效")
    if len(payload["review_brief"]) > MAX_REVIEW_BRIEF_LENGTH:
        raise ValueError("预览 Brief 超出长度限制")
    if not isinstance(payload["brief_summary"], str):
        raise ValueError("预览 Brief 摘要无效")

    rows = [_row_from_dict(row) for row in payload["rows"]]
    tests = [_test_from_dict(item) for item in payload["test_cases"]]
    row_numbers = [row.row_number for row in rows]
    if row_numbers != sorted(set(row_numbers)):
        raise ValueError("预览行号无效")
    if any(row.valid and not row.normalized.get("external_id") for row in rows):
        raise ValueError("预览内容编号无效")
    content_ids = [row.normalized.get("external_id") for row in rows if row.normalized.get("external_id")]
    duplicate_ids = {value for value in content_ids if content_ids.count(value) > 1}
    for duplicate in duplicate_ids:
        affected = [row for row in rows if row.normalized.get("external_id") == duplicate]
        if any(row.valid or not any("重复" in error for error in row.errors) for row in affected):
            raise ValueError("预览内容编号重复")
    test_ids = [test.external_test_case_id for test in tests]
    if any(not value for value in test_ids) or len(test_ids) != len(set(test_ids)):
        raise ValueError("预览测试场景编号无效或重复")
    content_id_set = set(content_ids)
    if any(not test.content_external_id or test.content_external_id not in content_id_set for test in tests):
        raise ValueError("预览测试场景引用不存在内容")
    for test in tests:
        if any(not _is_safe_basename(name) or Path(name).suffix.lower() not in EVIDENCE_SUFFIXES for name in test.evidence_filenames):
            raise ValueError("预览证据文件名无效")
    grouped = _group_tests(tests)
    for row in rows:
        expected = [_test_to_dict(test) for test in grouped.get(row.normalized.get("external_id") or "", [])]
        actual = [_test_to_dict(test) for test in row.tests]
        if actual != expected:
            raise ValueError("预览行测试场景绑定不匹配")

    valid_count = sum(row.valid for row in rows)
    expected_counts = (len(rows), valid_count, len(rows) - valid_count + len(payload["errors"]))
    stored_counts = (payload["total_count"], payload["valid_count"], payload["error_count"])
    if any(type(value) is not int or value < 0 for value in stored_counts):
        raise ValueError("预览计数类型无效")
    if stored_counts != expected_counts:
        raise ValueError("预览计数不匹配")

    return ImportPreview(
        token=payload["token"], rows=rows, warnings=list(payload["warnings"]),
        total_count=expected_counts[0], valid_count=expected_counts[1],
        error_count=expected_counts[2], review_brief=payload["review_brief"],
        brief_summary=payload["brief_summary"], test_cases=tests,
        errors=list(payload["errors"]), identity=_preview_identity_from_dict(payload),
    ), expires_at


def _row_from_dict(payload: Any) -> ImportRowPreview:
    if not isinstance(payload, dict) or set(payload) != _ROW_KEYS:
        raise ValueError("预览行字段无效")
    if type(payload["manuscript_index"]) is not int or payload["manuscript_index"] < 1:
        raise ValueError("预览稿件序号无效")
    if type(payload["row_number"]) is not int or payload["row_number"] < 2:
        raise ValueError("预览行号无效")
    normalized = payload["normalized"]
    legacy_normalized = frozenset(_COLUMN_KEYS[key] for key in IMPORT_COLUMNS)
    if not isinstance(normalized, dict) or set(normalized) not in {legacy_normalized, _NORMALIZED_KEYS}:
        raise ValueError("预览标准化字段无效")
    if any(value is not None and not isinstance(value, str) for value in normalized.values()):
        raise ValueError("预览标准化值无效")
    if not _is_string_list(payload["errors"]) or not _is_string_list(payload["warnings"]):
        raise ValueError("预览行消息无效")
    if type(payload["valid"]) is not bool or payload["valid"] != (not payload["errors"]):
        raise ValueError("预览行状态无效")
    if not isinstance(payload["tests"], list):
        raise ValueError("预览行测试场景无效")
    tests = [_test_from_dict(item) for item in payload["tests"]]
    return ImportRowPreview(
        manuscript_index=payload["manuscript_index"],
        row_number=payload["row_number"],
        normalized=dict(normalized),
        errors=list(payload["errors"]),
        warnings=list(payload["warnings"]),
        valid=payload["valid"],
        tests=tests,
    )


def _parse_manifest_expiry(value: Any) -> datetime:
    if not isinstance(value, str):
        raise ValueError("预览过期时间无效")
    expires_at = datetime.fromisoformat(value)
    if expires_at.tzinfo is None or expires_at.utcoffset() is None:
        raise ValueError("预览过期时间无时区")
    expires_at = expires_at.astimezone(timezone.utc)
    if expires_at > datetime.now(timezone.utc) + PREVIEW_TTL + timedelta(minutes=1):
        raise ValueError("预览过期时间超出限制")
    return expires_at


def _is_string_list(value: Any) -> bool:
    return isinstance(value, list) and all(isinstance(item, str) for item in value)


def cleanup_expired_previews() -> int:
    removed = 0
    now = datetime.now(timezone.utc)
    with _preview_locations_lock:
        for root in _load_preview_roots():
            if not root.is_dir():
                continue
            for preview_dir in root.iterdir():
                if not preview_dir.is_dir() or not _TOKEN_PATTERN.fullmatch(preview_dir.name):
                    continue
                manifest = preview_dir / "preview.json"
                try:
                    payload = json.loads(manifest.read_text(encoding="utf-8"))
                    preview, expires_at = _preview_from_dict(payload)
                    if preview.token != preview_dir.name or expires_at > now:
                        continue
                except (OSError, TypeError, ValueError, json.JSONDecodeError):
                    continue
                _preview_locations.pop(preview.token, None)
                shutil.rmtree(preview_dir, ignore_errors=True)
                removed += 1
    return removed
